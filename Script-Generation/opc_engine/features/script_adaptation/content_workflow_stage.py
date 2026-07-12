#!/usr/bin/env python3
import argparse
import csv
import io
import json
import os
import re
import shutil
import subprocess
import sys
import time
import warnings
from datetime import datetime
from pathlib import Path

from opc_engine.core.config_store import load_app_config
from opc_engine.core.project_assets import (
    ensure_project_dirs,
    infer_source_id,
    product_project_root,
    product_report_dir,
    raw_data_dir,
    require_product_project,
    source_stage_dir,
)
from opc_engine.features.video_teardown.analyze_video_teardown import (
    DEFAULT_BASE_URL,
    DEFAULT_MODEL,
    endpoint_variants,
    extract_text,
    get_api_key,
    merge_workflow_inputs,
    post_json,
)


ROOT = Path(__file__).resolve().parents[3]
DEFAULT_SCRIPT_ADAPTATION_PROMPT_PATH = ROOT / "workflow_configs" / "script_adaptation" / "config" / "script_adaptation_prompt.md"
DEFAULT_SCRIPT_ADAPTATION_PROMPT_CONFIG_PATH = "workflow_configs/script_adaptation/config/script_adaptation_prompt.md"
LEGACY_SCRIPT_ADAPTATION_PROMPT_PATH = ROOT / "knowledge_base" / "script_adaptation_prompt.md"
LEGACY_SCRIPT_ADAPTATION_PROMPT_CONFIG_PATH = "knowledge_base/script_adaptation_prompt.md"
DEFAULT_CONTENT_KNOWLEDGE_PATH = ROOT / "knowledge_base" / "hot_content_knowledge_base.md"
LEGACY_CONTENT_KNOWLEDGE_PATH = ROOT / "knowledge_base" / "video_teardown_knowledge_base.md"
DEFAULT_CONTENT_KNOWLEDGE_CONFIG_PATH = "knowledge_base/hot_content_knowledge_base.md"
LEGACY_CONTENT_KNOWLEDGE_CONFIG_PATH = "knowledge_base/video_teardown_knowledge_base.md"

METRICS_TABLE_SUFFIXES = {".csv", ".xlsx", ".xlsm"}


def log(message):
    print(message, flush=True)


def load_config(*workflow_stages):
    config = load_app_config()
    if workflow_stages and not os.environ.get("OPC_APP_CONFIG_PATH"):
        merge_workflow_inputs(config, *workflow_stages)
    return config


def resolve_path(value):
    raw_value = str(value or "").strip()
    if not raw_value:
        return None
    path = Path(raw_value).expanduser()
    if not path.is_absolute():
        path = ROOT / path
    return path.resolve()


def read_text(path):
    if not path or not path.exists() or not path.is_file():
        return ""
    return path.read_text(encoding="utf-8", errors="ignore").strip()


def resolve_project_path(value, default_path=None):
    raw_value = str(value or "").strip()
    if not raw_value and default_path:
        return default_path.resolve()
    path = Path(raw_value).expanduser()
    if not path.is_absolute():
        path = ROOT / path
    return path.resolve()


def normalize_content_knowledge_path(value):
    text = str(value or "").strip()
    if not text or text == LEGACY_CONTENT_KNOWLEDGE_CONFIG_PATH:
        return DEFAULT_CONTENT_KNOWLEDGE_CONFIG_PATH
    return text


def get_content_knowledge_base(config):
    knowledge_text = str(config.get("content_knowledge_base", "") or "").strip()
    if knowledge_text:
        return knowledge_text
    knowledge_path = resolve_project_path(
        normalize_content_knowledge_path(
            config.get("content_knowledge_base_path")
            or config.get("video_teardown_knowledge_base_path")
            or DEFAULT_CONTENT_KNOWLEDGE_CONFIG_PATH
        ),
        DEFAULT_CONTENT_KNOWLEDGE_PATH,
    )
    candidates = [knowledge_path]
    if knowledge_path != DEFAULT_CONTENT_KNOWLEDGE_PATH:
        candidates.append(DEFAULT_CONTENT_KNOWLEDGE_PATH)
    if LEGACY_CONTENT_KNOWLEDGE_PATH not in candidates:
        candidates.append(LEGACY_CONTENT_KNOWLEDGE_PATH)
    for candidate in candidates:
        if candidate.exists():
            return read_text(candidate)
    return ""


def get_script_adaptation_prompt(config):
    prompt = str(config.get("script_adaptation_prompt", "") or "").strip()
    if prompt:
        return prompt
    configured_path = str(config.get("script_adaptation_prompt_path") or "").strip()
    if not configured_path or configured_path == LEGACY_SCRIPT_ADAPTATION_PROMPT_CONFIG_PATH:
        configured_path = DEFAULT_SCRIPT_ADAPTATION_PROMPT_CONFIG_PATH
    prompt_path = resolve_project_path(
        configured_path,
        DEFAULT_SCRIPT_ADAPTATION_PROMPT_PATH,
    )
    if not prompt_path.exists() and LEGACY_SCRIPT_ADAPTATION_PROMPT_PATH.exists():
        return read_text(LEGACY_SCRIPT_ADAPTATION_PROMPT_PATH)
    return read_text(prompt_path)


def safe_name(value):
    text = str(value or "").strip() or "workflow"
    return "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in text).strip("_") or "workflow"


def timestamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def output_dir_for_stage(stage, config, anchor=""):
    if os.environ.get("OPC_APP_CONFIG_PATH"):
        unified_paths = config.get("unified_agent_paths") if isinstance(config.get("unified_agent_paths"), dict) else {}
        unified_key = {
            "adapt": "adapt_output_dir",
            "assemble": "assemble_output_dir",
        }.get(stage)
        if unified_key and unified_paths.get(unified_key):
            return Path(unified_paths[unified_key]).expanduser().resolve()
    ensure_project_dirs(config)
    anchor = str(anchor or "").strip()
    if stage == "metrics":
        return product_report_dir("data_attribution", config)
    if stage == "optimize":
        source_id = infer_source_id(anchor, "")
        if source_id:
            return source_stage_dir(source_id, "optimizations", config)
        return product_report_dir("script_optimizations", config)

    source_id = infer_source_id(anchor, "")
    source_stage = {
        "adapt": "adaptations",
        "assemble": "generated_videos",
        "publish": "publish_records",
    }.get(stage)
    if source_stage and source_id:
        return source_stage_dir(source_id, source_stage, config)

    fallback = {
        "adapt": "script_adaptations",
        "assemble": "generated_videos",
        "publish": "publish_records",
    }.get(stage, stage)
    return product_report_dir(fallback, config)


def write_outputs(stage, stem, markdown, payload, config=None, anchor=None, write_json=True):
    config = config or load_config()
    inferred_anchor = (
        anchor
        or payload.get("input_path")
        or payload.get("input_dir")
        or payload.get("video_path")
        or payload.get("script_path")
        or payload.get("metrics_path")
        or payload.get("merged_csv_path")
        or ""
    )
    output_dir = output_dir_for_stage(stage, config, inferred_anchor)
    output_dir.mkdir(parents=True, exist_ok=True)
    md_path = output_dir / f"{stem}.md"
    json_path = output_dir / f"{stem}.json"
    md_path.write_text(markdown.rstrip() + "\n", encoding="utf-8")
    log(f"Markdown 输出: {md_path}")
    if write_json:
        json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        log(f"JSON 输出: {json_path}")
    return md_path


def fenced_code_blocks(text, preferred_language=None):
    preferred = (preferred_language or "").strip().lower()
    blocks = []
    pattern = re.compile(r"```([A-Za-z0-9_-]*)\s*\n(.*?)```", re.DOTALL)
    for match in pattern.finditer(text or ""):
        language = (match.group(1) or "").strip().lower()
        body = match.group(2).strip()
        if preferred and language != preferred:
            continue
        blocks.append((language, body))
    return blocks


def strip_model_call_context(text):
    content = text or ""
    for marker in ["## 本次模型调用上下文", "# 本次模型调用上下文"]:
        if marker in content:
            return content.split(marker, 1)[0]
    return content


def parse_json_candidate(raw_text):
    text = (raw_text or "").strip()
    if not text:
        raise ValueError("empty json candidate")
    text = re.sub(r"^\s*json\s*", "", text, flags=re.IGNORECASE).strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        cleaned_lines = []
        for line in text.splitlines():
            stripped = line.strip()
            if stripped.startswith("//"):
                continue
            cleaned_lines.append(line)
        cleaned = "\n".join(cleaned_lines).strip()
        return json.loads(cleaned)


def balanced_json_candidates(text):
    candidates = []
    source = text or ""
    for start, char in enumerate(source):
        if char not in "{[":
            continue
        opener = char
        closer = "}" if opener == "{" else "]"
        stack = [closer]
        in_string = False
        escaped = False
        for pos in range(start + 1, len(source)):
            current = source[pos]
            if in_string:
                if escaped:
                    escaped = False
                elif current == "\\":
                    escaped = True
                elif current == '"':
                    in_string = False
                continue
            if current == '"':
                in_string = True
                continue
            if current in "{[":
                stack.append("}" if current == "{" else "]")
            elif stack and current == stack[-1]:
                stack.pop()
                if not stack:
                    candidates.append(source[start : pos + 1])
                    break
            elif current in "}]":
                break
    return candidates


def extract_image_prompt_json(adapted_text):
    adapted_text = strip_model_call_context(adapted_text)
    candidates = []
    candidates.extend(body for _, body in fenced_code_blocks(adapted_text, "json"))
    candidates.extend(
        body
        for language, body in fenced_code_blocks(adapted_text)
        if language != "json" and body.lstrip().startswith(("{", "["))
    )
    candidates.extend(balanced_json_candidates(adapted_text))

    ranked = sorted(
        candidates,
        key=lambda value: (
            "shots" not in value,
            "prompt_text" not in value,
            len(value),
        ),
    )
    last_error = None
    for candidate in ranked:
        try:
            parsed = parse_json_candidate(candidate)
        except Exception as exc:  # noqa: BLE001 - keep extraction resilient for model output.
            last_error = str(exc)
            continue
        if isinstance(parsed, dict) and ("shots" in parsed or "grid_layout" in parsed or "image_generation_model" in parsed):
            return parsed, ""
        if isinstance(parsed, list):
            return {"shots": parsed}, ""
    return None, last_error or "未找到模块一 JSON"


def csv_text_to_rows(csv_text):
    text = (csv_text or "").strip()
    if not text:
        raise ValueError("empty csv")
    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        raise ValueError("empty csv")
    header_index = 0
    for index, line in enumerate(lines):
        lower_line = line.lower()
        if "video_model_input_text" in lower_line or ("segment" in lower_line and "," in line):
            header_index = index
            break
    normalized_text = "\n".join(lines[header_index:]).strip()
    reader = csv.DictReader(io.StringIO(normalized_text))
    rows = [dict(row) for row in reader]
    fieldnames = list(reader.fieldnames or [])
    if not fieldnames or not rows:
        raise ValueError("CSV 表头或行内容为空")
    return normalized_text, fieldnames, rows


def parse_duration_seconds(value):
    text = str(value or "").strip().lower()
    if not text:
        return None
    text = text.replace("秒", "s").replace("ｓ", "s")
    text = re.sub(r"\s+", "", text)
    text = text[:-1] if text.endswith("s") else text
    if ":" in text:
        parts = text.split(":")
        try:
            seconds = 0.0
            for part in parts:
                seconds = seconds * 60 + float(part)
            return seconds
        except ValueError:
            return None
    try:
        return float(text)
    except ValueError:
        return None


def format_duration(seconds):
    rounded = round(float(seconds), 2)
    if rounded.is_integer():
        return f"{int(rounded)}s"
    return f"{rounded:.2f}".rstrip("0").rstrip(".") + "s"


def normalize_segment_duration(value):
    text = str(value or "").strip()
    if not text:
        return text
    parts = re.split(r"\s*(?:-|–|—|~|到|至)\s*", text, maxsplit=1)
    if len(parts) == 2:
        start = parse_duration_seconds(parts[0])
        end = parse_duration_seconds(parts[1])
        if start is not None and end is not None and end >= start:
            return format_duration(end - start)
    single = parse_duration_seconds(text)
    if single is not None and ("-" not in text and "–" not in text and "—" not in text and "~" not in text):
        return format_duration(single)
    return text


def normalize_shot_reference(value):
    text = str(value or "").strip()
    if not text:
        return text
    text = re.sub(r"\s*(?:[+/_-]\s*)?(?:continuation|continue|continued|延续片段|延续)\s*$", "", text, flags=re.IGNORECASE)
    return text.strip()


def normalize_video_prompt_row(row):
    normalized = dict(row)
    if "time_range" in normalized:
        normalized["time_range"] = normalize_segment_duration(normalized.get("time_range"))
    if "shot_reference" in normalized:
        normalized["shot_reference"] = normalize_shot_reference(normalized.get("shot_reference"))
    return normalized


def extract_markdown_segment_rows(adapted_text):
    text = adapted_text or ""
    marker_match = re.search(r"(模块二|分镜视频片段提示词|视频模型输入CSV|视频片段提示词)", text)
    if marker_match:
        text = text[marker_match.start() :]
    segment_pattern = re.compile(
        r"(?:\*\*)?【?\s*片段\s*([0-9A-Za-z_-]+)\s*[：:]\s*([^】\n*]+)\s*】?(?:\*\*)?",
        re.IGNORECASE,
    )
    matches = list(segment_pattern.finditer(text))
    rows = []
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        body = text[start:end].strip()
        body = re.sub(r"^[>*\-\s]+", "", body, flags=re.MULTILINE).strip()
        if not body:
            continue
        rows.append(
            {
                "segment_id": str(match.group(1)).strip(),
                "time_range": str(match.group(2)).strip(),
                "shot_reference": "",
                "video_model_input_text": re.sub(r"\s+", " ", body),
            }
        )
    return rows


def extract_video_prompt_rows(adapted_text):
    adapted_text = strip_model_call_context(adapted_text)
    csv_blocks = fenced_code_blocks(adapted_text, "csv")
    csv_blocks.extend(
        (language, body)
        for language, body in fenced_code_blocks(adapted_text)
        if language != "csv" and "," in body and "video_model_input_text" in body
    )
    ranked_blocks = sorted(
        [body for _, body in csv_blocks],
        key=lambda value: ("video_model_input_text" not in value.lower(), len(value)),
    )
    last_error = None
    for block in ranked_blocks:
        try:
            _, fieldnames, rows = csv_text_to_rows(block)
        except Exception as exc:  # noqa: BLE001 - model output can be messy.
            last_error = str(exc)
            continue
        if "video_model_input_text" not in fieldnames:
            last_error = "CSV 缺少 video_model_input_text 字段"
            continue
        return fieldnames, rows, ""

    fallback_rows = extract_markdown_segment_rows(adapted_text)
    if fallback_rows:
        return ["segment_id", "time_range", "shot_reference", "video_model_input_text"], fallback_rows, ""
    return [], [], last_error or "未找到模块二 CSV 或可解析片段"


def write_video_prompt_csv(path, fieldnames, rows):
    normalized_fields = list(fieldnames)
    for required in ["segment_id", "time_range", "shot_reference", "video_model_input_text"]:
        if required not in normalized_fields:
            normalized_fields.append(required)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=normalized_fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            normalized_row = normalize_video_prompt_row(row)
            writer.writerow({field: normalized_row.get(field, "") for field in normalized_fields})


def write_adaptation_structured_outputs(md_path, adapted_text, payload):
    image_data, image_error = extract_image_prompt_json(adapted_text)
    video_fieldnames, video_rows, video_error = extract_video_prompt_rows(adapted_text)
    outputs = {}
    errors = []

    if image_data is not None:
        image_path = md_path.with_name(f"{md_path.stem}_image_prompts.json")
        image_path.write_text(json.dumps(image_data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        outputs["image_prompts_json_path"] = str(image_path)
        outputs["image_prompt_count"] = len(image_data.get("shots", [])) if isinstance(image_data, dict) else 0
        log(f"文生图提示词 JSON: {image_path}")
    else:
        errors.append(f"模块一 JSON 提取失败：{image_error}")

    if video_rows:
        video_path = md_path.with_name(f"{md_path.stem}_video_prompts.csv")
        write_video_prompt_csv(video_path, video_fieldnames, video_rows)
        outputs["video_prompts_csv_path"] = str(video_path)
        outputs["video_prompt_count"] = len(video_rows)
        log(f"视频片段提示词 CSV: {video_path}")
    else:
        errors.append(f"模块二 CSV 提取失败：{video_error}")

    payload["structured_outputs"] = outputs
    if errors:
        payload["structured_output_errors"] = errors
    return errors


def split_script_into_segments(text, max_segments=12):
    if not text:
        return []
    blocks = []
    current = []
    for line in text.splitlines():
        if line.strip().startswith("镜头 ") and current:
            blocks.append("\n".join(current).strip())
            current = [line]
        else:
            current.append(line)
    if current:
        blocks.append("\n".join(current).strip())
    if len(blocks) <= 1:
        paragraphs = [part.strip() for part in text.split("\n\n") if part.strip()]
        blocks = paragraphs or [text.strip()]
    return blocks[:max_segments]


def build_local_adaptation_scaffold(input_path, source_text, target_model, segment_seconds, notes):
    segments = []
    for index, block in enumerate(split_script_into_segments(source_text), start=1):
        excerpt = block[:900]
        segments.append(
            {
                "segment_id": index,
                "duration_limit_seconds": segment_seconds,
                "source_excerpt": excerpt,
                "video_prompt_draft": f"将原脚本第 {index} 段改写为适合 {target_model} 生成的 {segment_seconds} 秒以内视频片段，保留动作、场景、产品展示与情绪节奏。",
                "first_frame_image_prompt": f"第 {index} 段首帧图：竖屏 9:16，真实 TikTok 原生感画面，人物/场景/产品状态与该片段开头动作一致，主体清晰，[product] 可见。",
            }
        )
    if not segments:
        segments.append(
            {
                "segment_id": 1,
                "duration_limit_seconds": segment_seconds,
                "source_excerpt": "",
                "video_prompt_draft": f"待填入成品脚本后，改写为适合 {target_model} 的 {segment_seconds} 秒以内视频片段。",
                "first_frame_image_prompt": "待填入成品脚本后，生成该片段首帧图描述。",
            }
        )

    lines = [
        "# 脚本适配结果",
        "",
        f"- 输入脚本：{input_path or '未选择'}",
        f"- 目标视频模型：{target_model}",
        f"- 单片段时长上限：{segment_seconds}s",
    ]
    if notes:
        lines.extend(["", "## 适配备注", notes])
    lines.extend(
        [
            "",
            "## 运行说明",
            "当前缺少成品脚本或脚本适配提示词，已生成本地占位框架。请补齐后重新运行，系统会输出可复制给大模型的完整适配提示词包。",
        ]
    )
    lines.extend(["", "## 片段适配框架"])
    for segment in segments:
        lines.extend(
            [
                "",
                f"### 片段 {segment['segment_id']}（≤ {segment_seconds}s）",
                "",
                "**原脚本参考：**",
                "",
                segment["source_excerpt"] or "待补充",
                "",
                "**视频生成提示词草案：**",
                "",
                segment["video_prompt_draft"],
                "",
                "**首帧图描述：**",
                "",
                segment["first_frame_image_prompt"],
            ]
        )
    return "\n".join(lines), segments


def build_adaptation_prompt(config, source_text, target_model, segment_seconds, notes):
    prompt_template = get_script_adaptation_prompt(config)
    content_knowledge = get_content_knowledge_base(config)
    return f"""# 系统自动导入变量（由页面结构化参数和本地文件生成）

目标视频生成模型：
{target_model}

单片段时长上限：
{segment_seconds}s

适配备注：
{notes or "无"}

爆款内容知识库：
{content_knowledge or "未填写爆款内容知识库。请严格参考脚本适配提示词，并提醒后续补充素材类型、原生感、转化逻辑和视频生成约束。"}

---

# 适配规则与输出格式提示词

{prompt_template}

---

# 本次执行要求

- 你正在做的是“脚本适配”：把已经产出的带货脚本改写成可直接服务于视频生成模型的分镜图提示词与视频片段生成指令。
- 必须同时参考“脚本适配提示词、成品脚本、爆款内容知识库”三类输入。
- 必须以系统注入的“目标视频生成模型”和“单片段时长上限”为准；不要把 Veo、8 秒或任何单一模型规则写死。
- 若目标模型有特殊能力或限制，请在适配结果中体现；若目标模型未知，采用通用保守的首帧图 + 视频生成提示词格式。
- 若脚本适配提示词与爆款内容知识库冲突，优先保证视频生成可执行性，再保留爆款内容知识库中的原生感、转化逻辑和素材框架。
- 不要输出分析过程，不要解释你怎么思考，直接输出可执行的适配结果。
- 产品视觉统一使用 [产品] 或 [手持产品]，不要虚构包装颜色、形状、材质或文字。

---

# 系统自动导入成品脚本

{source_text}
"""


def build_text_payload(prompt, max_output_tokens):
    return {
        "contents": [
            {
                "role": "user",
                "parts": [{"text": prompt}],
            }
        ],
        "generationConfig": {
            "temperature": 0.45,
            "maxOutputTokens": max_output_tokens,
        },
    }


def run_text_model(prompt, config, label):
    api_key = get_api_key(config)
    if not api_key:
        raise SystemExit("缺少 API Key：脚本适配需要调用文本模型，请先在视频拆解页保存 ModelMesh API Key，或设置 MODELMESH_API_KEY")

    model = (
        str(config.get("script_adaptation_text_model") or "").strip()
        or str(config.get("video_analysis_model") or "").strip()
        or DEFAULT_MODEL
    )
    base_url = str(config.get("modelmesh_base_url") or DEFAULT_BASE_URL).strip()
    max_output_tokens = int(config.get("video_analysis_max_output_tokens", 32768) or 32768)
    headers = {
        "Content-Type": "application/json",
        "x-goog-api-key": api_key,
    }

    log(f"开始调用文本模型完成{label}")
    log(f"文本模型: {model}")
    log(f"接口: {base_url.rstrip('/')}/v1beta/models/...:generateContent")

    last_error = None
    for url, endpoint_style in endpoint_variants(base_url, model):
        log(f"尝试接口格式: {endpoint_style}")
        payload = build_text_payload(prompt, max_output_tokens)
        status, response = post_json(url, headers, payload, 240)
        if 200 <= status < 300:
            return extract_text(response), response, endpoint_style
        last_error = {"status": status, "response": response, "endpoint_style": endpoint_style}
        message = response.get("error") if isinstance(response, dict) else response
        log(f"  未成功，HTTP {status}: {str(message)[:220]}")
        time.sleep(0.5)

    raise RuntimeError(f"{label}模型调用失败: {json.dumps(last_error, ensure_ascii=False)[:1200]}")


def run_adapt(config):
    input_path = resolve_path(config.get("script_adaptation_input_path"))
    source_text = read_text(input_path)
    target_model = str(config.get("script_adaptation_target_model") or "veo").strip()
    segment_seconds = int(config.get("script_adaptation_segment_seconds") or 8)
    notes = str(config.get("script_adaptation_notes") or "").strip()
    prompt_template = get_script_adaptation_prompt(config)
    knowledge_path = resolve_project_path(
        normalize_content_knowledge_path(
            config.get("content_knowledge_base_path")
            or config.get("video_teardown_knowledge_base_path")
            or DEFAULT_CONTENT_KNOWLEDGE_CONFIG_PATH
        ),
        DEFAULT_CONTENT_KNOWLEDGE_PATH,
    )

    log("开始脚本适配")
    log(f"目标视频生成模型: {target_model}")
    log(f"单片段时长上限: {segment_seconds}s")
    log(f"爆款内容知识库: {knowledge_path}")
    if input_path:
        log(f"输入脚本: {input_path}")
    else:
        log("未选择输入脚本，将生成空白适配框架")

    stem = f"{timestamp()}_{safe_name(input_path.stem if input_path else 'script_adaptation')}_{safe_name(target_model)}"
    payload = {
        "stage": "script_adaptation",
        "input_path": str(input_path) if input_path else "",
        "target_model": target_model,
        "segment_seconds": segment_seconds,
    }

    if source_text and prompt_template:
        adaptation_prompt = build_adaptation_prompt(config, source_text, target_model, segment_seconds, notes)
        adapted_text, raw_response, endpoint_style = run_text_model(adaptation_prompt, config, "脚本适配")
        segments = split_script_into_segments(source_text, max_segments=80)
        markdown = "\n".join(
            [
                "# 脚本适配结果",
                "",
                f"- 输入脚本：{input_path}",
                f"- 目标视频生成模型：{target_model}",
                f"- 单片段时长上限：{segment_seconds}s",
                f"- 爆款内容知识库：{knowledge_path}",
                f"- 本次处理：已调用文本模型完成适配；未调用视频生成模型。",
                f"- 接口格式：{endpoint_style}",
                "",
                "## 适配后内容",
                "",
                adapted_text.strip(),
                "",
                "## 本次模型调用上下文",
                "",
                "```text",
                adaptation_prompt.strip(),
                "```",
                "",
                "## 本地脚本切分参考",
                "",
                *(f"### 参考段 {index}\n\n{segment}" for index, segment in enumerate(segments, start=1)),
            ]
        )
        payload.update(
            {
                "adapted_text": adapted_text,
                "raw_response": raw_response,
                "adaptation_prompt": adaptation_prompt,
                "segment_count": len(segments),
                "segments": segments,
            }
        )
        md_path = write_outputs("adapt", stem, markdown, payload, config, write_json=False)
        extraction_errors = write_adaptation_structured_outputs(md_path, adapted_text, payload)
        if extraction_errors:
            raise RuntimeError("脚本适配完成，但结构化文件提取失败：" + "；".join(extraction_errors))
        log("脚本适配完成，已生成完整 Markdown、文生图 JSON 和视频片段 CSV")
        return

    if not source_text:
        log("未读取到成品脚本文本，回退为本地框架")
    if not prompt_template:
        log("未读取到脚本适配提示词，回退为本地框架")

    markdown, segments = build_local_adaptation_scaffold(input_path, source_text, target_model, segment_seconds, notes)
    payload["segments"] = segments

    write_outputs("adapt", stem, markdown, payload, config, write_json=False)
    log("脚本适配框架完成")


def run_assemble(config):
    input_dir = resolve_path(config.get("clip_assembly_input_dir"))
    output_name = safe_name(config.get("clip_assembly_output_name") or "assembled_video")
    notes = str(config.get("clip_assembly_notes") or "").strip()
    output_dir = output_dir_for_stage("assemble", config, input_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    log("开始视频生成流程框架")
    log("当前未接入 Veo/可灵等视频生成模型，也不是完整自动剪辑链路；仅处理已有片段目录、生成清单，必要时尝试本地合并。")
    log(f"片段目录: {input_dir or '未选择'}")
    clips = []
    if input_dir and input_dir.exists() and input_dir.is_dir():
        clips = sorted([p for p in input_dir.iterdir() if p.suffix.lower() in {".mp4", ".mov", ".m4v"}])
    log(f"检测到视频片段数量: {len(clips)}")

    manifest = {
        "stage": "clip_assembly",
        "input_dir": str(input_dir) if input_dir else "",
        "clips": [str(path) for path in clips],
        "notes": notes,
    }
    stem = f"{timestamp()}_{output_name}"
    manifest_path = output_dir / f"{stem}_manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    log(f"视频生成清单: {manifest_path}")

    ffmpeg = shutil.which("ffmpeg")
    if clips and ffmpeg:
        concat_list = output_dir / f"{stem}_concat.txt"
        concat_list.write_text("".join(f"file '{path.as_posix()}'\n" for path in clips), encoding="utf-8")
        output_video = output_dir / f"{stem}.mp4"
        log("检测到 ffmpeg，开始尝试无转码合并...")
        result = subprocess.run(
            [ffmpeg, "-y", "-f", "concat", "-safe", "0", "-i", str(concat_list), "-c", "copy", str(output_video)],
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
        )
        if result.returncode == 0 and output_video.exists():
            log(f"已有片段辅助合并完成: {output_video}")
            return
        log("无转码生成未成功，已保留视频生成清单，后续可改为转码合并。")
        log(result.stdout[-1200:])

    markdown = "\n".join(
        [
            "# 视频生成流程框架",
            "",
            "- 状态：流程框架，未接入实际视频生成模型和完整自动剪辑链路",
            f"- 片段目录：{input_dir or '未选择'}",
            f"- 检测片段：{len(clips)} 个",
            f"- 输出名称：{output_name}",
            f"- ffmpeg：{'已检测到' if ffmpeg else '未检测到'}",
            "",
            "## 片段顺序",
            *(f"{index}. {path.name}" for index, path in enumerate(clips, start=1)),
            "",
            "## 备注",
            notes or "待补充",
        ]
    )
    plan_path = output_dir / f"{stem}_plan.md"
    plan_path.write_text(markdown.rstrip() + "\n", encoding="utf-8")
    log(f"视频生成计划: {plan_path}")
    log("视频生成流程框架完成")


def run_publish(config):
    input_path = resolve_path(config.get("video_publish_input_path"))
    account = str(config.get("video_publish_account") or "").strip()
    caption = str(config.get("video_publish_caption") or "").strip()
    tags = str(config.get("video_publish_tags") or "").strip()
    mode = str(config.get("video_publish_mode") or "manual_record").strip()

    log("开始视频发布流程框架")
    log("当前未接入 TikTok 自动登录、账号授权和自动发布；仅生成本地发布计划/记录。")
    log(f"发布模式: {mode}")
    log(f"TikTok账号: {account or '未填写'}")
    log(f"视频文件: {input_path or '未选择'}")
    markdown = "\n".join(
        [
            "# 视频发布流程框架",
            "",
            f"- 状态：流程框架，未接入 TikTok 自动发布",
            f"- 发布状态：待发布",
            f"- 发布模式：{mode}",
            f"- TikTok账号：{account or '未填写'}",
            f"- 视频文件：{input_path or '未选择'}",
            "",
            "## 标题 / 文案",
            caption or "待填写",
            "",
            "## 标签",
            tags or "待填写",
            "",
            "## 接入说明",
            "当前为发布计划/记录框架，尚未接入 TikTok 自动发布。后续确认账号授权方式后，再改为自动发布任务。",
        ]
    )
    payload = {
        "stage": "video_publish",
        "status": "draft",
        "mode": mode,
        "account": account,
        "video_path": str(input_path) if input_path else "",
        "caption": caption,
        "tags": tags,
    }
    write_outputs("publish", f"{timestamp()}_{safe_name(account or 'publish_plan')}", markdown, payload, config)
    log("视频发布流程框架完成")


def parse_numeric(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").replace("%", "").strip())
    except ValueError:
        return None


def normalize_header(value):
    return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())


def find_column(headers, aliases):
    normalized = {normalize_header(header): header for header in headers}
    for alias in aliases:
        match = normalized.get(normalize_header(alias))
        if match:
            return match
    return ""


def normalize_video_id(value):
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    text = str(value).strip()
    if not text or text.upper() in {"N/A", "NA", "NONE", "-", "NULL"}:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def summarize_rows(rows, exclude_keys=None):
    exclude_tokens = [normalize_header(key) for key in (exclude_keys or [])]
    numeric_summary = {}
    if not rows:
        return numeric_summary
    keys = list(rows[0].keys())
    for key in keys:
        normalized_key = normalize_header(key)
        if any(token and token in normalized_key for token in exclude_tokens):
            continue
        values = [parse_numeric(row.get(key, "")) for row in rows]
        values = [value for value in values if value is not None]
        if values:
            numeric_summary[key] = {
                "count": len(values),
                "avg": round(sum(values) / len(values), 4),
                "sum": round(sum(values), 4),
            }
    return numeric_summary


def read_csv_rows(path):
    rows = []
    with path.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            rows.append(row)
    return rows


def read_xlsx_rows(path):
    from openpyxl import load_workbook

    rows = []
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    try:
        headers = next(iterator)
    except StopIteration:
        return rows, {}

    clean_headers = []
    for index, header in enumerate(headers, start=1):
        text = str(header or "").strip()
        clean_headers.append(text or f"column_{index}")

    for values in iterator:
        row = {}
        for header, value in zip(clean_headers, values):
            if hasattr(value, "isoformat"):
                row[header] = value.isoformat(sep=" ")
            elif value is None:
                row[header] = ""
            else:
                row[header] = value
        if any(str(value).strip() for value in row.values()):
            rows.append(row)
    workbook.close()
    return rows


def read_table_rows(path):
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path), "CSV"
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_rows(path), "Excel"
    return [], "Unsupported"


def summarize_table(path):
    rows, table_type = read_table_rows(path)
    return (rows, summarize_rows(rows)), table_type


NATURAL_VIDEO_ID_ALIASES = ["作品ID", "作品id", "视频ID", "视频id", "Video ID", "video id", "aweme id"]
ADS_VIDEO_ID_ALIASES = ["Video ID", "video id", "vidoe id", "视频ID", "视频id", "作品ID", "作品id"]
NATURAL_SIGNAL_ALIASES = ["播放量", "点赞量", "评论量", "账号用户名", "网页地址"]
ADS_SIGNAL_ALIASES = ["Cost", "SKU orders", "Gross revenue", "ROI", "Product ad impressions", "Product ad clicks"]

NATURAL_OUTPUT_COLUMNS = {
    "natural_user_id": ["用户ID", "User ID"],
    "natural_uid": ["UID"],
    "creator_name": ["账号用户名", "TikTok account", "达人", "creator"],
    "caption": ["文案", "Video title", "标题"],
    "cover_url": ["封面", "Cover"],
    "play_url": ["播放地址", "Play URL"],
    "tiktok_url": ["网页地址", "TikTok URL", "URL"],
    "organic_views": ["播放量", "Views"],
    "organic_likes": ["点赞量", "Likes"],
    "organic_comments": ["评论量", "Comments"],
    "created_at": ["创建时间", "发布时间", "Time posted"],
    "updated_at": ["更新时间"],
}

ADS_TEXT_COLUMNS = {
    "ad_video_title": ["Video title", "视频标题", "标题"],
    "ad_tiktok_account": ["TikTok account", "账号用户名", "达人"],
    "ad_creative_type": ["Creative type"],
    "ad_video_source": ["Video source"],
    "ad_status": ["Status"],
    "ad_time_posted": ["Time posted"],
    "currency": ["Currency"],
}

ADS_NUMERIC_SUM_COLUMNS = {
    "ad_cost": ["Cost", "Spend", "广告消耗", "花费"],
    "ad_sku_orders": ["SKU orders", "Orders", "订单", "成交"],
    "ad_gross_revenue": ["Gross revenue", "GMV", "Revenue", "销售额"],
    "ad_impressions": ["Product ad impressions", "Impressions", "曝光"],
    "ad_clicks": ["Product ad clicks", "Clicks", "点击"],
}

ADS_RATE_COLUMNS = {
    "ad_cost_per_order": ["Cost per order"],
    "ad_roi": ["ROI", "ROAS"],
    "ad_click_rate": ["Product ad click rate", "CTR"],
    "ad_conversion_rate": ["Ad conversion rate", "CVR"],
    "ad_2s_view_rate": ["2-second ad video view rate"],
    "ad_6s_view_rate": ["6-second ad video view rate"],
    "ad_25_view_rate": ["25% ad video view rate"],
    "ad_50_view_rate": ["50% ad video view rate"],
    "ad_75_view_rate": ["75% ad video view rate"],
    "ad_100_view_rate": ["100% ad video view rate"],
}

MERGED_METRICS_COLUMNS = [
    "video_id",
    "match_status",
    "creator_name",
    "caption",
    "tiktok_url",
    "play_url",
    "cover_url",
    "created_at",
    "updated_at",
    "organic_views",
    "organic_likes",
    "organic_comments",
    "ad_video_title",
    "ad_tiktok_account",
    "ad_creative_type",
    "ad_video_source",
    "ad_status",
    "ad_time_posted",
    "ad_cost",
    "ad_sku_orders",
    "ad_cost_per_order",
    "ad_gross_revenue",
    "ad_roi",
    "ad_impressions",
    "ad_clicks",
    "ad_click_rate",
    "ad_conversion_rate",
    "ad_2s_view_rate",
    "ad_6s_view_rate",
    "ad_25_view_rate",
    "ad_50_view_rate",
    "ad_75_view_rate",
    "ad_100_view_rate",
    "currency",
    "natural_user_id",
    "natural_uid",
]

MERGED_METRICS_COLUMN_LABELS = {
    "video_id": "视频ID",
    "match_status": "匹配状态",
    "creator_name": "创作者名称",
    "caption": "视频文案",
    "tiktok_url": "TikTok链接",
    "play_url": "播放地址",
    "cover_url": "封面地址",
    "created_at": "作品创建时间",
    "updated_at": "作品更新时间",
    "organic_views": "自然流累计播放量",
    "organic_likes": "自然流累计点赞量",
    "organic_comments": "自然流累计评论量",
    "ad_video_title": "广告视频标题",
    "ad_tiktok_account": "广告账号",
    "ad_creative_type": "广告创意类型",
    "ad_video_source": "广告视频来源",
    "ad_status": "广告状态",
    "ad_time_posted": "广告发布时间",
    "ad_cost": "广告消耗",
    "ad_sku_orders": "广告SKU订单数",
    "ad_cost_per_order": "广告单均成本",
    "ad_gross_revenue": "广告销售额",
    "ad_roi": "广告ROI/ROAS",
    "ad_impressions": "广告曝光量",
    "ad_clicks": "广告点击量",
    "ad_click_rate": "广告点击率",
    "ad_conversion_rate": "广告转化率",
    "ad_2s_view_rate": "2秒广告视频观看率",
    "ad_6s_view_rate": "6秒广告视频观看率",
    "ad_25_view_rate": "25%广告视频观看率",
    "ad_50_view_rate": "50%广告视频观看率",
    "ad_75_view_rate": "75%广告视频观看率",
    "ad_100_view_rate": "100%广告视频观看率",
    "currency": "币种",
    "natural_user_id": "自然流用户ID",
    "natural_uid": "自然流UID",
}

MATCH_STATUS_LABELS = {
    "matched": "两表已匹配",
    "natural_only": "仅自然流存在",
    "ads_only": "仅广告存在",
}


def detect_metrics_table_kind(rows):
    if not rows:
        return ""
    headers = list(rows[0].keys())
    natural_id = find_column(headers, NATURAL_VIDEO_ID_ALIASES)
    ads_id = find_column(headers, ADS_VIDEO_ID_ALIASES)
    natural_signals = sum(1 for alias in NATURAL_SIGNAL_ALIASES if find_column(headers, [alias]))
    ads_signals = sum(1 for alias in ADS_SIGNAL_ALIASES if find_column(headers, [alias]))
    if natural_id and natural_signals >= 2:
        return "natural"
    if ads_id and ads_signals >= 2:
        return "ads"
    return ""


def first_value(row, aliases):
    key = find_column(list(row.keys()), aliases)
    return row.get(key, "") if key else ""


def collect_unique_text(values):
    seen = []
    for value in values:
        text = str(value or "").strip()
        if text and text not in seen:
            seen.append(text)
    return " | ".join(seen[:5])


def average_numeric(values):
    parsed = [parse_numeric(value) for value in values]
    parsed = [value for value in parsed if value is not None]
    if not parsed:
        return ""
    return round(sum(parsed) / len(parsed), 6)


def aggregate_ads_rows(rows, ads_id_column):
    grouped = {}
    skipped_without_video_id = 0
    for row in rows:
        video_id = normalize_video_id(row.get(ads_id_column))
        if not video_id:
            skipped_without_video_id += 1
            continue
        grouped.setdefault(video_id, []).append(row)

    aggregated = {}
    for video_id, group_rows in grouped.items():
        item = {"video_id": video_id, "ad_row_count": len(group_rows)}
        for output_key, aliases in ADS_TEXT_COLUMNS.items():
            item[output_key] = collect_unique_text(first_value(row, aliases) for row in group_rows)
        for output_key, aliases in ADS_NUMERIC_SUM_COLUMNS.items():
            values = [parse_numeric(first_value(row, aliases)) for row in group_rows]
            values = [value for value in values if value is not None]
            item[output_key] = round(sum(values), 6) if values else ""
        for output_key, aliases in ADS_RATE_COLUMNS.items():
            item[output_key] = average_numeric(first_value(row, aliases) for row in group_rows)

        cost = parse_numeric(item.get("ad_cost"))
        orders = parse_numeric(item.get("ad_sku_orders"))
        revenue = parse_numeric(item.get("ad_gross_revenue"))
        impressions = parse_numeric(item.get("ad_impressions"))
        clicks = parse_numeric(item.get("ad_clicks"))
        if cost is not None and orders:
            item["ad_cost_per_order"] = round(cost / orders, 6)
        if cost:
            item["ad_roi"] = round((revenue or 0) / cost, 6)
        if impressions:
            item["ad_click_rate"] = round((clicks or 0) / impressions, 6)
        aggregated[video_id] = item
    return aggregated, skipped_without_video_id


def build_natural_output_row(row):
    item = {}
    for output_key, aliases in NATURAL_OUTPUT_COLUMNS.items():
        item[output_key] = first_value(row, aliases)
    return item


def merge_natural_and_ads(natural_rows, ads_rows):
    natural_headers = list(natural_rows[0].keys()) if natural_rows else []
    ads_headers = list(ads_rows[0].keys()) if ads_rows else []
    natural_id_column = find_column(natural_headers, NATURAL_VIDEO_ID_ALIASES)
    ads_id_column = find_column(ads_headers, ADS_VIDEO_ID_ALIASES)
    if not natural_id_column:
        raise ValueError("自然流数据缺少作品ID字段")
    if not ads_id_column:
        raise ValueError("投放数据缺少 Video ID 字段")

    ads_by_video_id, skipped_ads_without_video_id = aggregate_ads_rows(ads_rows, ads_id_column)
    natural_by_video_id = {}
    natural_order = []
    skipped_natural_without_video_id = 0
    for row in natural_rows:
        video_id = normalize_video_id(row.get(natural_id_column))
        if not video_id:
            skipped_natural_without_video_id += 1
            continue
        if video_id not in natural_by_video_id:
            natural_order.append(video_id)
        natural_by_video_id.setdefault(video_id, []).append(row)

    merged_rows = []
    matched = natural_only = ads_only = 0
    for video_id in natural_order:
        source_rows = natural_by_video_id[video_id]
        natural_item = build_natural_output_row(source_rows[0])
        ads_item = ads_by_video_id.get(video_id, {})
        match_status = "matched" if ads_item else "natural_only"
        matched += 1 if ads_item else 0
        natural_only += 0 if ads_item else 1
        merged_rows.append(
            {
                **{key: "" for key in MERGED_METRICS_COLUMNS},
                "video_id": video_id,
                "match_status": match_status,
                **natural_item,
                **ads_item,
            }
        )

    for video_id, ads_item in ads_by_video_id.items():
        if video_id in natural_by_video_id:
            continue
        ads_only += 1
        merged_rows.append(
            {
                **{key: "" for key in MERGED_METRICS_COLUMNS},
                "video_id": video_id,
                "match_status": "ads_only",
                **ads_item,
            }
        )

    summary = {
        "natural_rows": len(natural_rows),
        "ads_rows": len(ads_rows),
        "merged_rows": len(merged_rows),
        "matched_video_count": matched,
        "natural_only_video_count": natural_only,
        "ads_only_video_count": ads_only,
        "skipped_natural_without_video_id": skipped_natural_without_video_id,
        "skipped_ads_without_video_id": skipped_ads_without_video_id,
    }
    return merged_rows, summary


def write_metrics_table_outputs(stem, rows, config):
    output_dir = product_report_dir("data_attribution", config)
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / f"{stem}.csv"
    output_headers = [MERGED_METRICS_COLUMN_LABELS.get(column, column) for column in MERGED_METRICS_COLUMNS]
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=output_headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            output_row = {}
            for column in MERGED_METRICS_COLUMNS:
                label = MERGED_METRICS_COLUMN_LABELS.get(column, column)
                value = row.get(column, "")
                if column == "match_status":
                    value = MATCH_STATUS_LABELS.get(str(value), value)
                output_row[label] = value
            writer.writerow(output_row)
    log(f"归因 CSV 输出: {csv_path}")
    return csv_path


def localize_summary_keys(summary):
    localized = {}
    for key, value in summary.items():
        localized[MERGED_METRICS_COLUMN_LABELS.get(key, key)] = value
    return localized


def newest_metrics_table_in_dir(path, kind=""):
    if not path or not path.exists() or not path.is_dir():
        return None
    candidates = sorted(
        [item for item in path.rglob("*") if item.is_file() and item.suffix.lower() in METRICS_TABLE_SUFFIXES],
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    if not kind:
        return candidates[0] if candidates else None
    for candidate in candidates:
        try:
            rows, _ = read_table_rows(candidate)
            if detect_metrics_table_kind(rows) == kind:
                return candidate
        except Exception:
            continue
    return None


def resolve_metrics_table_path(config, config_key, kind="", fallback_dir=None):
    table_path = resolve_path(config.get(config_key))
    if table_path and table_path.is_dir():
        return newest_metrics_table_in_dir(table_path, kind)
    if table_path and table_path.exists():
        return table_path
    if fallback_dir:
        return newest_metrics_table_in_dir(fallback_dir, kind)
    return None


def normalize_entrypoint(value, default_value):
    text = str(value or "").strip()
    legacy_leaf = Path(text).name
    if legacy_leaf == "download_natural_flow_data.py":
        return "opc_engine.features.data_attribution.download_natural_flow_data"
    if legacy_leaf == "download_ad_performance_data.py":
        return "opc_engine.features.data_attribution.download_ad_performance_data"
    return text or default_value


def command_for_entrypoint(entrypoint):
    entrypoint = str(entrypoint or "").strip()
    if not entrypoint:
        return None
    module_name = entrypoint.removeprefix("module:").strip()
    if re.fullmatch(r"[A-Za-z_]\w*(\.[A-Za-z_]\w*)+", module_name):
        return [sys.executable, "-m", module_name]
    script_path = resolve_path(entrypoint)
    if not script_path or not script_path.exists() or not script_path.is_file():
        raise FileNotFoundError(f"内部下载入口不存在: {entrypoint}")
    suffix = script_path.suffix.lower()
    if suffix == ".py":
        return [sys.executable, str(script_path)]
    if suffix in {".sh", ".bash"}:
        return ["bash", str(script_path)]
    if os.access(script_path, os.X_OK):
        return [str(script_path)]
    return ["bash", str(script_path)]


def sanitize_download_log(text):
    replacements = {
        "GMVMax": "投放数据平台",
        "GMV Max": "投放数据平台",
        "TikTok Ads": "投放数据平台",
    }
    cleaned = str(text)
    for old, new in replacements.items():
        cleaned = cleaned.replace(old, new)
    cleaned = re.sub(r"https?://ads\.tiktok\.com[^\s)]+", "投放数据页面", cleaned)
    return cleaned


def run_download_script(config, label, script_key, default_entrypoint, output_subdir="", extra_env=None):
    entrypoint = normalize_entrypoint(config.get(script_key), default_entrypoint)
    output_dir = raw_data_dir(output_subdir or "downloads", config)
    notes = str(config.get("data_attribution_download_notes") or "").strip()
    output_dir.mkdir(parents=True, exist_ok=True)

    log(f"开始下载{label}")
    log(f"保存目录: {output_dir}")

    if not entrypoint:
        markdown = "\n".join(
            [
                f"# {label}下载",
                "",
                "- 状态：未配置内部下载入口",
                f"- 保存目录：{output_dir}",
                "",
                "## 下一步",
                "配置本地下载入口后再运行。",
                "",
                "## 备注",
                notes or "待补充",
            ]
        )
        write_outputs(
            "metrics",
            f"{timestamp()}_data_attribution_download_plan",
            markdown,
            {
                "stage": "data_attribution_download",
                "status": "script_not_configured",
                "label": label,
                "output_dir": str(output_dir),
                "notes": notes,
            },
            config,
        )
        log("未配置内部下载入口，已生成阶段一流程清单")
        return

    command = command_for_entrypoint(entrypoint)
    env = os.environ.copy()
    env["DATA_ATTRIBUTION_OUTPUT_DIR"] = str(output_dir)
    env["OPC_DATA_ATTRIBUTION_OUTPUT_DIR"] = str(output_dir)
    for key, value in (extra_env or {}).items():
        if value is not None and str(value).strip():
            env[key] = str(value)
    natural_url = config.get("natural_flow_management_url")
    natural_login_url = config.get("natural_flow_login_url")
    natural_group = config.get("natural_flow_account_group")
    natural_export_text = config.get("natural_flow_export_button_text_re")
    if natural_url:
        env["NATURAL_FLOW_MANAGEMENT_URL"] = str(natural_url)
    if natural_login_url:
        env["NATURAL_FLOW_LOGIN_URL"] = str(natural_login_url)
    if natural_group:
        env["NATURAL_FLOW_ACCOUNT_GROUP"] = str(natural_group)
        env["NATURAL_DATA_ACCOUNT_GROUP"] = str(natural_group)
    if natural_export_text:
        env["NATURAL_FLOW_EXPORT_TEXT_RE"] = str(natural_export_text)
        env["NATURAL_DATA_EXPORT_TEXT_RE"] = str(natural_export_text)
    log("执行内部下载流程...")
    process = subprocess.Popen(
        command,
        cwd=str(ROOT),
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        bufsize=1,
    )
    captured = []
    if process.stdout:
        for line in process.stdout:
            text = sanitize_download_log(line.rstrip())
            captured.append(text)
            log(text)
    exit_code = process.wait()
    downloaded_files = sorted([path for path in output_dir.rglob("*") if path.is_file()], key=lambda item: item.stat().st_mtime, reverse=True)

    markdown = "\n".join(
            [
                f"# {label}下载",
                "",
                f"- 状态：{'成功' if exit_code == 0 else '失败'}",
                f"- 退出码：{exit_code}",
                f"- 保存目录：{output_dir}",
                f"- 检测到文件：{len(downloaded_files)} 个",
                "",
                "## 最新文件",
            *(f"- {path}" for path in downloaded_files[:30]),
            "",
            "## 脚本输出摘要",
            "\n".join(captured[-80:]) or "无输出",
            "",
            "## 备注",
            notes or "待补充",
        ]
    )
    write_outputs(
        "metrics",
        f"{timestamp()}_{safe_name(label)}_download",
        markdown,
        {
            "stage": "data_attribution_download",
            "status": "success" if exit_code == 0 else "failed",
            "exit_code": exit_code,
            "label": label,
            "output_dir": str(output_dir),
            "downloaded_files": [str(path) for path in downloaded_files],
            "notes": notes,
        },
        config,
    )
    if exit_code != 0:
        raise SystemExit(exit_code)
    log(f"{label}下载完成")


def run_metrics_natural_download(config):
    run_download_script(
        config,
        "自然流数据",
        "data_attribution_download_script_path",
        "opc_engine.features.data_attribution.download_natural_flow_data",
        output_subdir="natural_flow",
    )


def run_metrics_ads_download(config):
    run_download_script(
        config,
        "投放数据",
        "data_attribution_ads_download_script_path",
        "opc_engine.features.data_attribution.download_ad_performance_data",
        output_subdir="ad_performance",
        extra_env={"AD_PERFORMANCE_DATE_RANGE": "yesterday"},
    )


def run_metrics_download(config):
    run_metrics_natural_download(config)
    run_metrics_ads_download(config)


def run_metrics(config):
    configured_download_dir = str(config.get("data_attribution_download_output_dir") or "").strip()
    if configured_download_dir == "metrics/raw_downloads":
        download_output_dir = product_project_root(config) / "raw_data"
    else:
        download_output_dir = resolve_path(configured_download_dir) or (product_project_root(config) / "raw_data")
    legacy_input_path = resolve_path(config.get("data_recovery_input_path"))
    natural_input_path = resolve_metrics_table_path(
        config,
        "data_recovery_natural_input_path",
        "natural",
        download_output_dir,
    )
    ads_input_path = resolve_metrics_table_path(config, "data_recovery_ads_input_path", "ads", download_output_dir)
    if legacy_input_path and legacy_input_path.is_dir():
        natural_input_path = natural_input_path or newest_metrics_table_in_dir(legacy_input_path, "natural")
        ads_input_path = ads_input_path or newest_metrics_table_in_dir(legacy_input_path, "ads")
        input_path = natural_input_path or ads_input_path or newest_metrics_table_in_dir(legacy_input_path)
    else:
        input_path = legacy_input_path
    manual_metrics = str(config.get("data_recovery_manual_metrics") or "").strip()
    if not input_path and not (natural_input_path and ads_input_path) and download_output_dir:
        input_path = newest_metrics_table_in_dir(download_output_dir)

    log("开始数据归因阶段二：整理与分析")
    log(f"自然流数据: {'已找到最新原始表' if natural_input_path else '未找到'}")
    log(f"投放数据: {'已找到最新原始表' if ads_input_path else '未找到'}")

    if natural_input_path and ads_input_path:
        natural_rows, natural_table_type = read_table_rows(natural_input_path)
        ads_rows, ads_table_type = read_table_rows(ads_input_path)
        log(f"读取自然流 {natural_table_type} 行数: {len(natural_rows)}")
        log(f"读取投放 {ads_table_type} 行数: {len(ads_rows)}")
        merged_rows, attribution_summary = merge_natural_and_ads(natural_rows, ads_rows)
        metrics_summary = summarize_rows(
            merged_rows,
            exclude_keys=["id", "url", "path", "time", "created", "updated", "caption", "title", "status", "type", "currency"],
        )
        localized_metrics_summary = localize_summary_keys(metrics_summary)
        stem = f"{timestamp()}_作品归因汇总"
        csv_path = write_metrics_table_outputs(stem, merged_rows, config)
        markdown = "\n".join(
            [
                "# 数据归因结果",
                "",
                "阶段二已按同一个作品维度，把自然流数据和投放数据合并成一张作品归因表。",
                "",
                f"- 合并 CSV：{csv_path}",
                "",
                "## 匹配结果",
                f"- 自然流原始行数：{attribution_summary['natural_rows']}",
                f"- 投放原始行数：{attribution_summary['ads_rows']}",
                f"- 合并后作品数：{attribution_summary['merged_rows']}",
                f"- 两表匹配作品数：{attribution_summary['matched_video_count']}",
                f"- 仅自然流存在：{attribution_summary['natural_only_video_count']}",
                f"- 仅投放存在：{attribution_summary['ads_only_video_count']}",
                f"- 自然流无作品ID跳过：{attribution_summary['skipped_natural_without_video_id']}",
                f"- 投放无 Video ID 跳过：{attribution_summary['skipped_ads_without_video_id']}",
                "",
                "## 手动数据",
                manual_metrics or "待补充",
                "",
                "## 数值字段汇总",
                json.dumps(localized_metrics_summary, ensure_ascii=False, indent=2) if localized_metrics_summary else "暂无可汇总数值字段",
            ]
        )
        payload = {
            "stage": "data_attribution",
            "mode": "natural_ads_merge",
            "natural_input_detected": bool(natural_input_path),
            "ads_input_detected": bool(ads_input_path),
            "merged_csv_path": str(csv_path),
            "manual_metrics": manual_metrics,
            "attribution_summary": attribution_summary,
            "numeric_summary": localized_metrics_summary,
        }
        write_outputs("metrics", stem, markdown, payload, config)
        log("数据归因阶段二完成")
        return

    log(f"单表数据来源: {'已找到最新原始表' if input_path else '手动填写/未选择'}")

    rows = []
    summary = {}
    table_type = ""
    if input_path and input_path.exists() and input_path.suffix.lower() in METRICS_TABLE_SUFFIXES:
        (rows, summary), table_type = summarize_table(input_path)
        log(f"读取 {table_type} 行数: {len(rows)}")
    markdown = "\n".join(
        [
            "# 数据归因结果",
            "",
            f"- 数据来源：{'已读取最新原始表' if input_path else '手动填写/未选择'}",
            f"- 数据类型：{table_type or '手动/未识别'}",
            f"- 数据行数：{len(rows)}",
            "",
            "## 手动数据",
            manual_metrics or "待补充",
            "",
            "## 数值字段汇总",
            json.dumps(summary, ensure_ascii=False, indent=2) if summary else "暂无可汇总数值字段",
        ]
    )
    payload = {
        "stage": "data_attribution",
        "input_path": str(input_path) if input_path else "",
        "table_type": table_type,
        "manual_metrics": manual_metrics,
        "row_count": len(rows),
        "numeric_summary": summary,
    }
    write_outputs("metrics", f"{timestamp()}_{safe_name(input_path.stem if input_path else 'metrics')}", markdown, payload, config)
    log("数据归因阶段二完成")


def run_optimize(config):
    script_path = resolve_path(config.get("script_optimization_input_path"))
    metrics_path = resolve_path(config.get("script_optimization_metrics_path"))
    notes = str(config.get("script_optimization_notes") or "").strip()
    source_script = read_text(script_path)
    metrics_text = read_text(metrics_path)

    log("开始脚本优化框架")
    log(f"原脚本: {script_path or '未选择'}")
    log(f"数据文件: {metrics_path or '未选择'}")
    markdown = "\n".join(
        [
            "# 脚本优化建议",
            "",
            f"- 原脚本：{script_path or '未选择'}",
            f"- 数据文件：{metrics_path or '未选择'}",
            "",
            "## 加权评估框架",
            "- 播放完成/停留：判断 Hook 与前 3 秒是否成立。",
            "- 点击/互动：判断冲突、痛点、评论诱因是否成立。",
            "- 转化/GMV：判断产品机制、信任背书、价格锚点是否成立。",
            "- 多视频加权平均：后续接入发布数据后，以视频级指标反推脚本表现。",
            "",
            "## 当前优化建议草案",
            "1. 先定位数据最弱的环节：开头停留、互动、点击、成交。",
            "2. 保留表现强的镜头结构，只替换低效话术和弱视觉证据。",
            "3. 对同一脚本拆出 A/B 版本：强冲突版、强证明版、强价格锚点版。",
            "",
            "## 备注",
            notes or "待补充",
            "",
            "## 原脚本摘要",
            source_script[:1600] or "待补充",
            "",
            "## 数据摘要",
            metrics_text[:1600] or "待补充",
        ]
    )
    payload = {
        "stage": "script_optimization",
        "script_path": str(script_path) if script_path else "",
        "metrics_path": str(metrics_path) if metrics_path else "",
        "notes": notes,
    }
    write_outputs("optimize", f"{timestamp()}_{safe_name(script_path.stem if script_path else 'script_optimization')}", markdown, payload, config)
    log("脚本优化框架完成")


def parse_args():
    parser = argparse.ArgumentParser(description="Run local content distribution workflow scaffolds.")
    parser.add_argument(
        "stage",
        choices=["adapt", "assemble", "publish", "metrics_download", "metrics_natural_download", "metrics_ads_download", "metrics", "optimize"],
        help="要运行的工作流阶段",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    stage_inputs = {
        "adapt": ("script_adaptation",),
        "assemble": ("video_generation",),
        "publish": ("video_publish",),
        "metrics_download": ("data_attribution",),
        "metrics_natural_download": ("data_attribution",),
        "metrics_ads_download": ("data_attribution",),
        "metrics": ("data_attribution",),
        "optimize": ("script_optimization",),
    }
    config = load_config(*stage_inputs.get(args.stage, ()))
    if not os.environ.get("OPC_APP_CONFIG_PATH"):
        require_product_project(config, "执行内容工作流")
        ensure_project_dirs(config)
    {
        "adapt": run_adapt,
        "assemble": run_assemble,
        "publish": run_publish,
        "metrics_download": run_metrics_download,
        "metrics_natural_download": run_metrics_natural_download,
        "metrics_ads_download": run_metrics_ads_download,
        "metrics": run_metrics,
        "optimize": run_optimize,
    }[args.stage](config)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        log(f"任务失败: {exc}")
        sys.exit(1)
