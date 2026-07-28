#!/usr/bin/env python3
import argparse
import csv
import io
import json
import math
import os
import re
import shutil
import subprocess
import sys
import time
import warnings
from datetime import datetime
from pathlib import Path

from opc_engine.core.config_store import load_app_config
from opc_engine.core.project_assets import (
    ensure_project_dirs,
    infer_source_id,
    product_project_root,
    product_report_dir,
    raw_data_dir,
    require_product_project,
    source_stage_dir,
)
from opc_engine.features.video_teardown.analyze_video_teardown import (
    DEFAULT_BASE_URL,
    DEFAULT_MODEL,
    endpoint_variants,
    extract_text,
    get_api_key,
    merge_workflow_inputs,
    post_json,
)


ROOT = Path(__file__).resolve().parents[3]
DEFAULT_SCRIPT_ADAPTATION_PROMPT_PATH = ROOT / "workflow_configs" / "script_adaptation" / "config" / "script_adaptation_prompt.md"
DEFAULT_SCRIPT_ADAPTATION_PROMPT_CONFIG_PATH = "workflow_configs/script_adaptation/config/script_adaptation_prompt.md"
LEGACY_SCRIPT_ADAPTATION_PROMPT_PATH = ROOT / "knowledge_base" / "script_adaptation_prompt.md"
LEGACY_SCRIPT_ADAPTATION_PROMPT_CONFIG_PATH = "knowledge_base/script_adaptation_prompt.md"
METRICS_TABLE_SUFFIXES = {".csv", ".xlsx", ".xlsm"}


def log(message):
    print(message, flush=True)


def load_config(*workflow_stages):
    config = load_app_config()
    if workflow_stages:
        merge_workflow_inputs(config, *workflow_stages)
    return config


def resolve_path(value):
    raw_value = str(value or "").strip()
    if not raw_value:
        return None
    path = Path(raw_value).expanduser()
    if not path.is_absolute():
        path = ROOT / path
    return path.resolve()


def read_text(path):
    if not path or not path.exists() or not path.is_file():
        return ""
    return path.read_text(encoding="utf-8", errors="ignore").strip()


def resolve_project_path(value, default_path=None):
    raw_value = str(value or "").strip()
    if not raw_value and default_path:
        return default_path.resolve()
    path = Path(raw_value).expanduser()
    if not path.is_absolute():
        path = ROOT / path
    return path.resolve()


def get_script_adaptation_prompt(config):
    prompt = str(config.get("script_adaptation_prompt", "") or "").strip()
    if prompt:
        return prompt
    target_model = str(config.get("script_adaptation_target_model") or "veo").strip()
    prompt_paths = config.get("script_adaptation_prompt_paths")
    configured_path = ""
    if isinstance(prompt_paths, dict):
        configured_path = str(prompt_paths.get(target_model) or "").strip()
    if not configured_path:
        configured_path = str(config.get("script_adaptation_prompt_path") or "").strip()
    if not configured_path or configured_path == LEGACY_SCRIPT_ADAPTATION_PROMPT_CONFIG_PATH:
        configured_path = DEFAULT_SCRIPT_ADAPTATION_PROMPT_CONFIG_PATH
    prompt_path = resolve_project_path(
        configured_path,
        DEFAULT_SCRIPT_ADAPTATION_PROMPT_PATH,
    )
    if not prompt_path.exists() and LEGACY_SCRIPT_ADAPTATION_PROMPT_PATH.exists():
        return read_text(LEGACY_SCRIPT_ADAPTATION_PROMPT_PATH)
    return read_text(prompt_path)


def safe_name(value):
    text = str(value or "").strip() or "workflow"
    return "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in text).strip("_") or "workflow"


def parse_script_filename(value):
    stem = Path(str(value or "")).stem
    parts = [part for part in stem.split("-") if part]
    if parts and parts[0].lower() in {"veo", "omni", "grok", "kling", "hailuo", "runway"} and len(parts) > 1:
        parts = parts[1:]

    result = {
        "stem": stem,
        "adaptation_type": "",
        "product_name": "",
        "country": "",
        "username": "",
        "video_id": "",
        "variant_index": "",
        "has_country_format": False,
    }
    if len(parts) < 3 or parts[0] not in {"复刻", "裂变"}:
        return result

    result["adaptation_type"] = parts[0]
    country_index = None
    for index in range(2, max(len(parts) - 1, 2)):
        token = parts[index].strip()
        if re.fullmatch(r"[A-Z]{2,3}", token):
            country_index = index
            break

    if country_index is not None and country_index + 2 < len(parts) + 1:
        result["product_name"] = "-".join(parts[1:country_index])
        result["country"] = parts[country_index]
        result["username"] = parts[country_index + 1] if country_index + 1 < len(parts) else ""
        video_part = "-".join(parts[country_index + 2 :])
        result["has_country_format"] = True
    else:
        result["product_name"] = parts[1] if len(parts) > 1 else ""
        result["username"] = parts[2] if len(parts) > 2 else ""
        video_part = "-".join(parts[3:]) if len(parts) > 3 else ""

    match = re.fullmatch(r"(.+?)_(\d{2,})", video_part)
    if match:
        result["video_id"] = match.group(1)
        result["variant_index"] = match.group(2)
    else:
        result["video_id"] = video_part
    return result


def resolve_optional_path(value):
    raw_value = str(value or "").strip()
    if not raw_value:
        return None
    path = Path(raw_value).expanduser()
    if not path.is_absolute():
        path = ROOT / path
    return path.resolve()


def script_adaptation_output_root(config):
    target_model = str(config.get("script_adaptation_target_model") or "veo").strip()
    output_roots = config.get("script_adaptation_output_roots")
    if isinstance(output_roots, dict):
        target_root = output_roots.get(target_model)
        if target_root:
            return resolve_optional_path(target_root)
    return resolve_optional_path(config.get("script_adaptation_output_root"))


def script_adaptation_input_dir(config):
    return resolve_optional_path(config.get("script_adaptation_input_dir"))


def product_folder_from_script(config, anchor=""):
    configured = str(config.get("script_adaptation_product_folder") or "").strip()
    if configured:
        return safe_name(configured)

    name = Path(str(anchor or "")).stem
    parsed = parse_script_filename(name)
    if parsed.get("product_name"):
        return safe_name(parsed["product_name"])

    input_root = script_adaptation_input_dir(config)
    anchor_path = resolve_optional_path(anchor)
    if input_root and anchor_path:
        try:
            relative = anchor_path.relative_to(input_root)
            if len(relative.parts) > 1:
                return safe_name(relative.parts[0])
        except ValueError:
            pass

    for separator in ("-", "_"):
        if separator in name:
            candidate = name.split(separator, 1)[0]
            if candidate:
                return safe_name(candidate)
    return safe_name(name or "未分类")


def timestamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def output_dir_for_stage(stage, config, anchor=""):
    ensure_project_dirs(config)
    anchor = str(anchor or "").strip()
    if stage == "adapt":
        external_root = script_adaptation_output_root(config)
        if external_root:
            relative_dir = Path(str(config.get("script_adaptation_output_relative_dir") or "").strip())
            if str(relative_dir) not in {"", "."} and not relative_dir.is_absolute() and ".." not in relative_dir.parts:
                return external_root / relative_dir
            return external_root / product_folder_from_script(config, anchor)
    if stage == "metrics":
        return product_report_dir("data_attribution", config)
    if stage == "optimize":
        source_id = infer_source_id(anchor, "")
        if source_id:
            return source_stage_dir(source_id, "optimizations", config)
        return product_report_dir("script_optimizations", config)

    source_id = infer_source_id(anchor, "")
    source_stage = {
        "adapt": "adaptations",
        "assemble": "generated_videos",
        "publish": "publish_records",
    }.get(stage)
    if source_stage and source_id:
        return source_stage_dir(source_id, source_stage, config)

    fallback = {
        "adapt": "script_adaptations",
        "assemble": "generated_videos",
        "publish": "publish_records",
    }.get(stage, stage)
    return product_report_dir(fallback, config)


def write_outputs(stage, stem, markdown, payload, config=None, anchor=None, write_json=True):
    config = config or load_config()
    inferred_anchor = (
        anchor
        or payload.get("input_path")
        or payload.get("input_dir")
        or payload.get("video_path")
        or payload.get("script_path")
        or payload.get("metrics_path")
        or payload.get("merged_csv_path")
        or ""
    )
    output_dir = output_dir_for_stage(stage, config, inferred_anchor)
    output_dir.mkdir(parents=True, exist_ok=True)
    md_path = output_dir / f"{stem}.md"
    json_path = output_dir / f"{stem}.json"
    md_path.write_text(markdown.rstrip() + "\n", encoding="utf-8")
    log(f"Markdown 输出: {md_path}")
    if write_json:
        json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        log(f"JSON 输出: {json_path}")
    return md_path


def fenced_code_blocks(text, preferred_language=None):
    preferred = (preferred_language or "").strip().lower()
    blocks = []
    pattern = re.compile(r"```([A-Za-z0-9_-]*)\s*\n(.*?)```", re.DOTALL)
    for match in pattern.finditer(text or ""):
        language = (match.group(1) or "").strip().lower()
        body = match.group(2).strip()
        if preferred and language != preferred:
            continue
        blocks.append((language, body))
    return blocks


def strip_model_call_context(text):
    content = text or ""
    for marker in ["## 本次模型调用上下文", "# 本次模型调用上下文"]:
        if marker in content:
            return content.split(marker, 1)[0]
    return content


def parse_json_candidate(raw_text):
    text = (raw_text or "").strip()
    if not text:
        raise ValueError("empty json candidate")
    text = re.sub(r"^\s*json\s*", "", text, flags=re.IGNORECASE).strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        cleaned_lines = []
        for line in text.splitlines():
            stripped = line.strip()
            if stripped.startswith("//"):
                continue
            cleaned_lines.append(line)
        cleaned = "\n".join(cleaned_lines).strip()
        return json.loads(cleaned)


def balanced_json_candidates(text):
    candidates = []
    source = text or ""
    for start, char in enumerate(source):
        if char not in "{[":
            continue
        opener = char
        closer = "}" if opener == "{" else "]"
        stack = [closer]
        in_string = False
        escaped = False
        for pos in range(start + 1, len(source)):
            current = source[pos]
            if in_string:
                if escaped:
                    escaped = False
                elif current == "\\":
                    escaped = True
                elif current == '"':
                    in_string = False
                continue
            if current == '"':
                in_string = True
                continue
            if current in "{[":
                stack.append("}" if current == "{" else "]")
            elif stack and current == stack[-1]:
                stack.pop()
                if not stack:
                    candidates.append(source[start : pos + 1])
                    break
            elif current in "}]":
                break
    return candidates


def extract_image_prompt_json(adapted_text):
    adapted_text = strip_model_call_context(adapted_text)
    candidates = []
    candidates.extend(body for _, body in fenced_code_blocks(adapted_text, "json"))
    candidates.extend(
        body
        for language, body in fenced_code_blocks(adapted_text)
        if language != "json" and body.lstrip().startswith(("{", "["))
    )
    candidates.extend(balanced_json_candidates(adapted_text))

    ranked = sorted(
        candidates,
        key=lambda value: (
            "shots" not in value,
            "prompt_text" not in value,
            len(value),
        ),
    )
    last_error = None
    for candidate in ranked:
        try:
            parsed = parse_json_candidate(candidate)
        except Exception as exc:  # noqa: BLE001 - keep extraction resilient for model output.
            last_error = str(exc)
            continue
        if isinstance(parsed, dict) and ("shots" in parsed or "grid_layout" in parsed or "image_generation_model" in parsed):
            return parsed, ""
        if isinstance(parsed, list):
            return {"shots": parsed}, ""
    return None, last_error or "未找到模块一 JSON"


def expected_grid_layout(shot_count):
    if shot_count <= 4:
        return "2x2"
    if shot_count <= 6:
        return "2x3"
    if shot_count <= 9:
        return "3x3"
    return "3x3"


def normalize_storyboard_grid_json(image_data):
    if not isinstance(image_data, dict):
        return image_data
    normalized = dict(image_data)
    shots = normalized.get("shots")
    if not isinstance(shots, list):
        shots = []
        normalized["shots"] = shots
    shot_count = len(shots)
    normalized["image_generation_model"] = normalized.get("image_generation_model") or "NanoBananaPro"
    normalized["output_mode"] = "storyboard_grid_preview"
    normalized["grid_layout"] = expected_grid_layout(shot_count)
    normalized["allowed_grid_layouts"] = ["2x2", "2x3", "3x3"]
    normalized["grid_aspect_ratio"] = "9:16"
    normalized["grid_rendering_rules"] = {
        "canvas": "single clean 9:16 storyboard grid image",
        "cell_geometry": "all cells must be perfectly equal width and equal height, straight borders, no perspective distortion",
        "layout": f"strict {normalized['grid_layout']} grid, left-to-right and top-to-bottom order",
        "gutters": "thin consistent white gutters between cells, outer margin consistent on all sides",
        "blank_cells": "unused cells must be plain white or very light neutral empty cells with no objects, no people, no product, no text, no numbers",
        "visual_flow_cutting": "designed for deterministic equal-grid cropping; do not create collage, overlapping panels, tilted frames, irregular cells, rounded panels, labels, captions, or decorative borders",
    }
    normalized["meta"] = {
        "valid_shots": shot_count,
        "crop_order": list(range(1, shot_count + 1)),
    }
    normalized["blank_cell_policy"] = {
        "enabled": True,
        "style": "plain white or very light neutral empty cell",
        "no_objects": True,
        "no_people": True,
        "no_product": True,
        "no_text": True,
        "no_number": True,
    }
    normalized["export_rules"] = {
        "split_grid": True,
        "export_only_real_shots": True,
        "skip_blank_cells": True,
        "expected_export_count": shot_count,
    }
    required_suffix = "clean image, no stickers, no subtitles, no timecode, no visible text, no numbers"
    for index, shot in enumerate(shots, start=1):
        if not isinstance(shot, dict):
            continue
        shot["shot_number"] = f"{index:02d}"
        prompt_text = str(shot.get("prompt_text") or "").strip()
        for token in required_suffix.split(", "):
            if token not in prompt_text:
                prompt_text = f"{prompt_text}, {token}" if prompt_text else token
        shot["prompt_text"] = prompt_text
    return normalized


def remove_first_storyboard_json(text):
    content = strip_model_call_context(text or "")
    pattern = re.compile(r"```(?:json)?\s*\n(.*?)\n```", re.DOTALL | re.IGNORECASE)
    for match in pattern.finditer(content):
        try:
            parsed = parse_json_candidate(match.group(1))
        except Exception:
            continue
        if isinstance(parsed, dict) and ("shots" in parsed or parsed.get("output_mode") == "storyboard_grid_preview"):
            return (content[: match.start()] + content[match.end() :]).strip()
    return content.strip()


def strip_outer_markdown_fence(text):
    content = strip_model_call_context(text or "").strip()
    match = re.fullmatch(r"```(?:markdown|md)?\s*\n(.*?)\n```", content, re.DOTALL | re.IGNORECASE)
    if match:
        return match.group(1).strip() + "\n"
    return content.rstrip() + "\n" if content else ""


OMNI_OMITTED_SCRIPT_FIELDS = {"背景音乐", "字幕", "贴纸", "特效"}
OMNI_SCRIPT_MARKER_PATTERN = re.compile(r"下面是(?:我的完整脚本|本段镜头脚本(?:（已过滤字段）)?)\s*[：:]\s*")
OMNI_SEGMENT_HEADING_PATTERN = re.compile(
    r"(?m)^(?:#{1,6}\s*)?(?:\*\*)?镜头\s*(?P<number>\d+)"
    r"(?:\s*-\s*part\s*\d+)?(?:\*\*)?\s*[（(]\s*"
    r"(?P<start>\d+(?::\d{1,2}){0,2}(?:\.\d+)?)\s*(?:-|–|—|~|至|到)\s*"
    r"(?P<end>\d+(?::\d{1,2}){0,2}(?:\.\d+)?)[^）)\n]*[）)](?:\*\*)?[^\n]*$",
    re.IGNORECASE,
)


def remove_chinese_translation_notes(text):
    content = str(text or "")
    content = re.sub(r"[（(][^）)]*中文翻译对照[：:][^）)]*[）)]", "", content)
    content = re.sub(r"[（(][^）)]*中文翻译[：:][^）)]*[）)]", "", content)
    content = re.sub(r"[；;，,]\s*中文翻译对照[：:][^）)]*", "", content)
    content = re.sub(r"[；;，,]\s*中文翻译[：:][^）)]*", "", content)
    content = re.sub(r"[（(]\s*中文翻译对照[：:][^）)]*[）)]", "", content)
    content = re.sub(r"[（(]\s*中文翻译[：:][^）)]*[）)]", "", content)
    return content


def parse_omni_time_value(value):
    text = str(value or "").strip()
    if not text:
        return None
    try:
        parts = [float(part) for part in text.split(":")]
    except ValueError:
        return None
    if len(parts) == 1:
        return parts[0]
    if len(parts) == 2:
        return parts[0] * 60 + parts[1]
    if len(parts) == 3:
        return parts[0] * 3600 + parts[1] * 60 + parts[2]
    return None


def format_omni_time(seconds):
    seconds = max(float(seconds or 0), 0.0)
    minutes = int(seconds // 60)
    remainder = seconds - minutes * 60
    return f"{minutes:02d}:{remainder:06.3f}"


def split_omni_segments(text):
    content = str(text or "")
    matches = list(re.finditer(r"(?m)^#\s*Segment\s+.*$", content))
    if not matches:
        return [("", content)]
    segments = []
    if matches[0].start() > 0:
        segments.append(("", content[: matches[0].start()]))
    for index, match in enumerate(matches):
        end = matches[index + 1].start() if index + 1 < len(matches) else len(content)
        segments.append(("segment", content[match.start() : end]))
    return segments


def reset_omni_segment_embedded_script(block):
    marker = OMNI_SCRIPT_MARKER_PATTERN.search(block)
    if not marker:
        return block
    prefix = block[: marker.end()]
    script = block[marker.end() :]
    matches = list(OMNI_SEGMENT_HEADING_PATTERN.finditer(script))
    if not matches:
        return block

    parsed_ranges = []
    for match in matches:
        start = parse_omni_time_value(match.group("start"))
        end = parse_omni_time_value(match.group("end"))
        if start is None or end is None or end < start:
            return block
        parsed_ranges.append((start, end))
    base_start = parsed_ranges[0][0]
    local_end = parsed_ranges[-1][1] - base_start
    local_shots = "、".join(f"镜头{index}" for index in range(1, len(matches) + 1))
    local_range = f"{format_omni_time(0)} - {format_omni_time(local_end)}"

    segment_title = re.search(r"(?m)^#\s*Segment\s+([0-9A-Za-z_-]+)", prefix)
    if segment_title:
        prefix = re.sub(
            r"(?m)^#\s*Segment\s+.*$",
            f"# Segment {segment_title.group(1)}：{local_range}",
            prefix,
            count=1,
        )
    prefix = re.sub(
        r"本段包含[^\n。]*。请严格",
        f"本段包含{local_shots}。请严格",
        prefix,
    )
    prefix = re.sub(
        r"按脚本时间轴展示【?\d+(?::\d{1,2}){0,2}(?:\.\d+)?\s*(?:-|–|—|~|至|到)\s*"
        r"\d+(?::\d{1,2}){0,2}(?:\.\d+)?】?内",
        f"按本段本地时间轴展示{local_range}内",
        prefix,
    )
    prefix = prefix.replace("按脚本时间轴展示", "按本段本地时间轴展示")

    def replace_heading(match):
        index = len(replace_heading.ranges) + 1
        start, end = parsed_ranges[index - 1]
        replace_heading.ranges.append((start, end))
        local_start = start - base_start
        local_stop = end - base_start
        return f"### 镜头 {index} ({format_omni_time(local_start)} - {format_omni_time(local_stop)})"

    replace_heading.ranges = []
    return prefix + OMNI_SEGMENT_HEADING_PATTERN.sub(replace_heading, script)


def reset_omni_output_segment_scripts(text):
    parts = []
    for kind, block in split_omni_segments(text):
        parts.append(reset_omni_segment_embedded_script(block) if kind == "segment" else block)
    return "".join(parts)


def omni_embedded_script_reset_issues(text):
    issues = []
    segment_index = 0
    for kind, block in split_omni_segments(text):
        if kind != "segment":
            continue
        segment_index += 1
        marker = OMNI_SCRIPT_MARKER_PATTERN.search(block)
        if not marker:
            continue
        matches = list(OMNI_SEGMENT_HEADING_PATTERN.finditer(block[marker.end() :]))
        if not matches:
            issues.append(f"Segment {segment_index} 本段镜头脚本缺少镜头标题")
            continue
        for local_index, match in enumerate(matches, start=1):
            number = int(match.group("number"))
            start = parse_omni_time_value(match.group("start"))
            if number != local_index:
                issues.append(f"Segment {segment_index} 镜头编号未重置：第 {local_index} 个写成镜头 {number}")
                break
            if local_index == 1 and (start is None or abs(start) > 0.01):
                issues.append(f"Segment {segment_index} 起始时间未重置为 00:00.000")
                break
    return issues


def sanitize_omni_script_text(text):
    lines = remove_chinese_translation_notes(text).splitlines()
    cleaned = []
    skipping_field = None
    field_pattern = re.compile(r"^\s*[*-]\s*(?:\*\*)?\[([^\]]+)\](?:\*\*)?")
    for line in lines:
        field_match = field_pattern.match(line)
        if field_match:
            field_name = re.sub(r"[（(].*$", "", field_match.group(1).strip()).strip()
            if field_name in OMNI_OMITTED_SCRIPT_FIELDS:
                skipping_field = field_name
                continue
            skipping_field = None
        elif skipping_field:
            starts_next_block = (
                line.startswith("#")
                or line.startswith("---")
                or re.match(r"^\s*[*-]\s*(?:\*\*)?\[[^\]]+\](?:\*\*)?", line)
            )
            if not starts_next_block:
                continue
            skipping_field = None
        cleaned.append(line)
    return "\n".join(cleaned).strip()


def neutralize_omni_product_descriptions(text):
    content = str(text or "")
    content = re.sub(
        r"(?m)^4\.\s*与产品使用相关的手部姿势参考.*$",
        "4. 中性手部与上半身自然姿态参考；不得手持产品，不得出现产品，不得做产品使用动作。",
        content,
    )
    content = re.sub(
        r"(?m)^5\.\s*如有必要，展示角色使用产品前后的状态差异.*$",
        "5. 如需表现状态差异，只展示人物自身发色、胡须或皮肤状态差异；不得出现产品、产品包装、产品道具或使用动作。",
        content,
    )
    content = content.replace("使用产品手势参考", "中性手部姿态参考")
    content = content.replace("手持该产品状产品", "自然手部姿态")
    content = content.replace("手持该产品产品", "手持图1产品")
    content = re.sub(
        r"左上角区域\(产品与角色设定\):展示图1产品图中的.*?，以及搭配该产品的图2人物半身参考图和关联道具。",
        "左上角区域(产品与角色设定):展示图1中的该产品，以及搭配该产品的图2人物半身参考图和关联道具。",
        content,
    )
    content = re.sub(
        r"产品外观必须参考图1，不要改变产品颜色、结构、包装、品牌视觉和关键细节。",
        "该产品外观只由图1决定；提示词不得描述或推断该产品的品类、颜色、结构、包装、品牌文字、材质、形状和细节。",
        content,
    )
    product_rule = "动作不能改变图1产品外观。"
    anchor = "该产品外观只由图1决定；提示词不得描述或推断该产品的品类、颜色、结构、包装、品牌文字、材质、形状和细节。"
    if product_rule not in content and anchor in content:
        content = content.replace(anchor, f"{anchor}\n{product_rule}", 1)
    content = re.sub(
        r"(?i)\b(?:Stylo\s*Hair\s*Color|SIMC)\b",
        "该产品",
        content,
    )
    action_replacements = {
        "向后脑勺方向梳理": "沿发丝方向移动",
        "向后脑勺梳理": "沿发丝方向移动",
        "梳理头发和胡须": "进行遮白动作",
        "梳理头发": "沿发丝移动",
        "梳理头顶": "沿头顶发丝移动",
        "梳理胡须": "沿胡须区域移动",
        "梳理": "沿发丝移动",
        "梳过": "经过",
        "划过": "经过",
        "刷过": "经过",
        "刮": "经过",
        "涂抹": "进行遮白动作",
        "涂满了黑色的该产品": "完成遮白效果",
        "湿润的黑色该产品质感": "自然遮白后的发丝质感",
        "黑色该产品质感": "自然遮白后的发丝质感",
        "整齐的黑色印记": "清晰的遮白效果",
        "黑色的痕迹": "清晰的遮白效果",
        "黑色痕迹": "清晰的遮白效果",
        "染成黑色": "完成遮白",
        "染黑": "完成遮白",
        "全黑年轻容貌": "自然遮白后的精神状态",
        "全部变成整齐的黑色": "呈现自然遮白效果",
        "变成浓密的黑色": "呈现自然遮白效果",
        "变黑": "完成遮白",
        "被均匀的黑色覆盖": "呈现均匀遮白效果",
        "黑色该产品": "遮白效果",
        "黑色产品": "遮白效果",
        "黑色膏体": "遮白效果",
        "露出该产品": "进行使用演示",
        "旋出该产品": "进行使用演示",
        "冲洗掉头发上的该产品": "冲洗头发",
        "水流冲经过": "水流冲刷",
        "由上至下沿发丝移动": "由上至下沿胡须区域移动",
    }
    for old, new in action_replacements.items():
        content = content.replace(old, new)
    content = re.sub(r"`?\[产品\]`?", "图1产品", content)
    content = re.sub(r"手持`?\[产品\]`?", "手持图1产品", content)
    content = re.sub(r"被`?\[产品\]`?经过", "图1产品经过", content)
    content = re.sub(r"在`?\[产品\]`?[^，。；\n]*?下", "经过图1产品动作后", content)
    content = re.sub(r"从该产品中[^，。；\n]*", "进行使用演示", content)
    content = re.sub(r"该产品质地[^，。；\n]*", "遮白效果自然", content)
    content = re.sub(r"该产品[^，。；\n]*?排列[^，。；\n]*", "动作清晰", content)
    replacements = {
        "梳状的染发产品": "该产品",
        "梳状染发产品": "该产品",
        "梳状产品": "该产品",
        "黑色梳状产品": "该产品",
        "染发梳产品": "该产品",
        "染发棒产品": "该产品",
        "染发产品": "该产品",
        "染发梳": "该产品",
        "染发棒": "该产品",
        "梳子": "该产品",
        "刷头": "该产品",
        "产品包装盒": "该产品",
        "[产品]包装盒": "[产品]",
        "`[产品]`包装盒": "`[产品]`",
        "包装盒": "该产品",
        "瓶身": "该产品",
        "梳齿结构": "该产品",
        "梳齿缝隙": "该产品",
        "细密梳齿": "该产品",
        "梳齿": "该产品",
        "旋钮": "该产品",
        "深黑色的膏体": "该产品",
        "黑色的膏体": "该产品",
        "染发膏": "该产品",
        "膏体": "该产品",
        "产品使用部位": "该产品",
        "产品操作部位": "该产品",
        "产品内容物": "该产品",
        "产品细节": "该产品",
    }
    for old, new in replacements.items():
        content = content.replace(old, new)
    content = re.sub(r"手持[^，。；\n]*?该产品", "手持图1产品", content)
    content = re.sub(r"图1产品图中的[^，。\n]*?该产品", "图1中的该产品", content)
    return content


def sanitize_omni_output_markdown(text):
    return neutralize_omni_product_descriptions(
        reset_omni_output_segment_scripts(sanitize_omni_script_text(text))
    ).strip()


def omni_source_duration_seconds(text):
    ranges = []
    for match in OMNI_SEGMENT_HEADING_PATTERN.finditer(str(text or "")):
        start = parse_omni_time_value(match.group("start"))
        end = parse_omni_time_value(match.group("end"))
        if start is not None and end is not None and end >= start:
            ranges.append((start, end))
    if not ranges:
        return None
    start = min(item[0] for item in ranges)
    end = max(item[1] for item in ranges)
    return end - start if end > start else None


def expected_omni_segment_count(source_text, segment_seconds=10):
    duration = omni_source_duration_seconds(source_text)
    if duration is None:
        return None, None
    seconds = max(float(segment_seconds or 10), 1.0)
    if 30.0 <= duration <= 40.0 and abs(seconds - 10.0) < 0.01:
        return 4, duration
    return max(1, int(math.ceil((duration - 0.001) / seconds))), duration


def fixed_duration_padding_plan(source_text, segment_seconds):
    duration = omni_source_duration_seconds(source_text)
    if duration is None:
        return None
    seconds = max(float(segment_seconds or 1), 1.0)
    remainder = duration % seconds
    if remainder < 0.001 or seconds - remainder < 0.001:
        remainder = seconds
        padding = 0.0
    else:
        padding = seconds - remainder
    return {
        "source_duration": duration,
        "segment_seconds": seconds,
        "last_content_duration": remainder,
        "padding_duration": padding,
    }


def fixed_duration_padding_note(source_text, target_model, segment_seconds):
    target_key = str(target_model or "").strip().lower()
    if target_key not in {"omni", "veo"}:
        return ""
    plan = fixed_duration_padding_plan(source_text, segment_seconds)
    if not plan:
        return ""
    if plan["padding_duration"] < 0.001:
        return (
            f"\n{str(target_model).title()} 原始时长保真要求：\n"
            f"原脚本有效总时长为 {plan['source_duration']:.3f}s，恰好填满固定时长片段。"
            "不得延长、缩短或新增技术占位。\n"
        )
    return (
        f"\n{str(target_model).title()} 固定时长技术占位硬性要求：\n"
        f"原脚本有效总时长为 {plan['source_duration']:.3f}s；"
        f"最后一个 Segment 只有 {plan['last_content_duration']:.3f}s 原始内容，"
        f"必须在尾部增加 {plan['padding_duration']:.3f}s 黑屏静音技术占位，"
        f"使该 Segment 达到 {plan['segment_seconds']:.3f}s。\n"
        "不得通过慢动作、延长停留、重复动作、补充台词、增加空镜或新增剧情来填满时长。\n"
        "技术占位不属于原脚本镜头，不得拆成新的 Segment；必须原样输出以下机器标记：\n"
        f"- 原脚本总时长：{plan['source_duration']:.3f}秒\n"
        f"- 本段有效内容时长：{plan['last_content_duration']:.3f}秒\n"
        f"- 有效内容结束：{plan['last_content_duration']:.3f}秒\n"
        f"- 技术占位开始：{plan['last_content_duration']:.3f}秒\n"
        f"- 技术占位时长：{plan['padding_duration']:.3f}秒\n"
        f"- 模型片段时长：{plan['segment_seconds']:.3f}秒\n"
        "[TECHNICAL_PADDING: BLACK_SILENT]\n"
        "占位区必须描述为纯黑画面、完全静音、无人物、无产品、无字幕、无贴纸、无动作、无转场内容。\n"
    )


def normalize_omni_machine_structure(markdown, source_text, segment_seconds):
    content = str(markdown or "").strip()
    full_required = (
        "AI视频生成分段提示词包",
        "## 1. 分段总览",
        "## 2. 人物设定总览",
        "## 3. 每段生成提示词",
    )
    if "## 每段生成提示词" not in content and not all(item in content for item in full_required):
        content = f"#\n## 每段生成提示词\n\n---\n\n{content}"

    plan = fixed_duration_padding_plan(source_text, segment_seconds)
    if not plan or plan["padding_duration"] < 0.001:
        return content.strip()

    field_labels = (
        "原脚本总时长",
        "本段有效内容时长",
        "有效内容结束",
        "技术占位开始",
        "技术占位时长",
        "模型片段时长",
    )
    field_pattern = "|".join(re.escape(label) for label in field_labels)
    content = re.sub(
        rf"(?m)^[ \t]*-?[ \t]*(?:{field_pattern})[：:].*(?:\n|$)",
        "",
        content,
    )
    content = re.sub(
        r"(?m)^[ \t]*\[TECHNICAL_PADDING: BLACK_SILENT\][ \t]*(?:\n|$)",
        "",
        content,
    )
    canonical_description = "技术占位为纯黑画面、完全静音、无人物、无产品、无字幕、无贴纸、无动作、无转场内容。"
    content = re.sub(
        rf"(?m)^[ \t]*{re.escape(canonical_description)}[ \t]*(?:\n|$)",
        "",
        content,
    )

    segment_headings = list(re.finditer(r"(?m)^#\s*Segment\s+.+$", content))
    if not segment_headings:
        return content.strip()

    machine_block = (
        f"\n- 原脚本总时长：{plan['source_duration']:.3f}秒\n"
        f"- 本段有效内容时长：{plan['last_content_duration']:.3f}秒\n"
        f"- 有效内容结束：{plan['last_content_duration']:.3f}秒\n"
        f"- 技术占位开始：{plan['last_content_duration']:.3f}秒\n"
        f"- 技术占位时长：{plan['padding_duration']:.3f}秒\n"
        f"- 模型片段时长：{plan['segment_seconds']:.3f}秒\n"
        "[TECHNICAL_PADDING: BLACK_SILENT]\n"
        f"{canonical_description}\n"
    )
    insert_at = segment_headings[-1].end()
    return f"{content[:insert_at]}{machine_block}{content[insert_at:]}".strip()


def omni_segment_count_issues(output_text, source_text="", segment_seconds=10):
    expected_count, duration = expected_omni_segment_count(source_text, segment_seconds)
    if expected_count is None:
        return []
    actual_count = len(re.findall(r"(?m)^#\s*Segment\s+.+$", str(output_text or "")))
    if actual_count > expected_count:
        return [
            f"Omni 分段数量过多：原视频约 {duration:.1f}s，最多应为 {expected_count} 段，当前 {actual_count} 段"
        ]
    return []


def build_clean_adaptation_markdown(input_path, target_model, segment_seconds, endpoint_style, adapted_text):
    image_data, image_error = extract_image_prompt_json(adapted_text)
    body_text = remove_first_storyboard_json(adapted_text)
    parts = [
        "# 脚本适配结果",
        "",
        f"- 输入脚本：{input_path}",
        f"- 目标视频生成模型：{target_model}",
        f"- 单片段时长上限：{segment_seconds}s",
        "- 本次处理：已调用文本模型完成适配；未调用视频生成模型。",
        f"- 接口格式：{endpoint_style}",
        "",
    ]
    if image_data:
        normalized = normalize_storyboard_grid_json(image_data)
        parts.extend(
            [
                "## 模块一：宫格分镜 JSON",
                "",
                "```json",
                json.dumps(normalized, ensure_ascii=False, indent=2),
                "```",
                "",
            ]
        )
    else:
        parts.extend(
            [
                "## 模块一：宫格分镜 JSON",
                "",
                f"> 未能提取可校正的宫格 JSON：{image_error}",
                "",
            ]
        )
    if body_text:
        parts.extend(["## 模块二：Veo 线性实操手册", "", body_text])
    return "\n".join(parts).rstrip() + "\n"


def csv_text_to_rows(csv_text):
    text = (csv_text or "").strip()
    if not text:
        raise ValueError("empty csv")
    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        raise ValueError("empty csv")
    header_index = 0
    for index, line in enumerate(lines):
        lower_line = line.lower()
        if "video_model_input_text" in lower_line or ("segment" in lower_line and "," in line):
            header_index = index
            break
    normalized_text = "\n".join(lines[header_index:]).strip()
    reader = csv.DictReader(io.StringIO(normalized_text))
    rows = [dict(row) for row in reader]
    fieldnames = list(reader.fieldnames or [])
    if not fieldnames or not rows:
        raise ValueError("CSV 表头或行内容为空")
    return normalized_text, fieldnames, rows


def parse_duration_seconds(value):
    text = str(value or "").strip().lower()
    if not text:
        return None
    text = text.replace("秒", "s").replace("ｓ", "s")
    text = re.sub(r"\s+", "", text)
    text = text[:-1] if text.endswith("s") else text
    if ":" in text:
        parts = text.split(":")
        try:
            seconds = 0.0
            for part in parts:
                seconds = seconds * 60 + float(part)
            return seconds
        except ValueError:
            return None
    try:
        return float(text)
    except ValueError:
        return None


def format_duration(seconds):
    rounded = round(float(seconds), 2)
    if rounded.is_integer():
        return f"{int(rounded)}s"
    return f"{rounded:.2f}".rstrip("0").rstrip(".") + "s"


def normalize_segment_duration(value):
    text = str(value or "").strip()
    if not text:
        return text
    parts = re.split(r"\s*(?:-|–|—|~|到|至)\s*", text, maxsplit=1)
    if len(parts) == 2:
        start = parse_duration_seconds(parts[0])
        end = parse_duration_seconds(parts[1])
        if start is not None and end is not None and end >= start:
            return format_duration(end - start)
    single = parse_duration_seconds(text)
    if single is not None and ("-" not in text and "–" not in text and "—" not in text and "~" not in text):
        return format_duration(single)
    return text


def normalize_shot_reference(value):
    text = str(value or "").strip()
    if not text:
        return text
    text = re.sub(r"\s*(?:[+/_-]\s*)?(?:continuation|continue|continued|延续片段|延续)\s*$", "", text, flags=re.IGNORECASE)
    return text.strip()


def normalize_video_prompt_row(row):
    normalized = dict(row)
    if "time_range" in normalized:
        normalized["time_range"] = normalize_segment_duration(normalized.get("time_range"))
    if "shot_reference" in normalized:
        normalized["shot_reference"] = normalize_shot_reference(normalized.get("shot_reference"))
    return normalized


VEO_BATCH_CSV_FIELDS = ["序号", "提示词", "横竖屏", "模型系列", "清晰度", "图片模式", "首帧图片", "尾帧图片", "参考图", "次数"]


def longest_digit_sequence(value):
    matches = re.findall(r"\d+", str(value or ""))
    return max(matches, key=len) if matches else ""


def reference_image_id_from_payload(payload, md_path):
    candidates = [
        payload.get("input_path", ""),
        payload.get("output_stem", ""),
        md_path.stem,
    ]
    matches = [longest_digit_sequence(candidate) for candidate in candidates]
    matches = [match for match in matches if match]
    return max(matches, key=len) if matches else ""


def extract_markdown_segment_rows(adapted_text):
    text = adapted_text or ""
    marker_match = re.search(r"(模块二|分镜视频片段提示词|视频模型输入CSV|视频片段提示词)", text)
    if marker_match:
        text = text[marker_match.start() :]
    segment_pattern = re.compile(
        r"(?:\*\*)?【?\s*片段\s*([0-9A-Za-z_-]+)\s*[：:]\s*([^】\n*]+)\s*】?(?:\*\*)?",
        re.IGNORECASE,
    )
    matches = list(segment_pattern.finditer(text))
    rows = []
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        body = text[start:end].strip()
        body = re.sub(r"^[>*\-\s]+", "", body, flags=re.MULTILINE).strip()
        if not body:
            continue
        rows.append(
            {
                "segment_id": str(match.group(1)).strip(),
                "time_range": str(match.group(2)).strip(),
                "shot_reference": "",
                "video_model_input_text": re.sub(r"\s+", " ", body),
            }
        )
    return rows


def extract_video_prompt_rows(adapted_text):
    adapted_text = strip_model_call_context(adapted_text)
    csv_blocks = fenced_code_blocks(adapted_text, "csv")
    csv_blocks.extend(
        (language, body)
        for language, body in fenced_code_blocks(adapted_text)
        if language != "csv" and "," in body and "video_model_input_text" in body
    )
    ranked_blocks = sorted(
        [body for _, body in csv_blocks],
        key=lambda value: ("video_model_input_text" not in value.lower(), len(value)),
    )
    last_error = None
    for block in ranked_blocks:
        try:
            _, fieldnames, rows = csv_text_to_rows(block)
        except Exception as exc:  # noqa: BLE001 - model output can be messy.
            last_error = str(exc)
            continue
        if "video_model_input_text" not in fieldnames:
            last_error = "CSV 缺少 video_model_input_text 字段"
            continue
        return fieldnames, rows, ""

    fallback_rows = extract_markdown_segment_rows(adapted_text)
    if fallback_rows:
        return ["segment_id", "time_range", "shot_reference", "video_model_input_text"], fallback_rows, ""
    return [], [], last_error or "未找到模块二 CSV 或可解析片段"


def video_prompt_text_from_row(row):
    return (
        row.get("video_model_input_text")
        or row.get("提示词")
        or row.get("prompt")
        or row.get("Prompt")
        or ""
    )


def write_video_prompt_csv(path, fieldnames, rows, reference_image_id=""):
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=VEO_BATCH_CSV_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for index, row in enumerate(rows, start=1):
            normalized_row = normalize_video_prompt_row(row)
            writer.writerow(
                {
                    "序号": index,
                    "提示词": video_prompt_text_from_row(normalized_row),
                    "横竖屏": 1,
                    "模型系列": 1,
                    "清晰度": 1,
                    "图片模式": 2,
                    "首帧图片": "",
                    "尾帧图片": "",
                    "参考图": reference_image_id,
                    "次数": 1,
                }
            )


def write_adaptation_structured_outputs(md_path, adapted_text, payload):
    image_data, image_error = extract_image_prompt_json(adapted_text)
    video_fieldnames, video_rows, video_error = extract_video_prompt_rows(adapted_text)
    outputs = {}
    errors = []

    if image_data is not None:
        image_path = md_path.with_name(f"{md_path.stem}_image_prompts.json")
        image_path.write_text(json.dumps(image_data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        outputs["image_prompts_json_path"] = str(image_path)
        outputs["image_prompt_count"] = len(image_data.get("shots", [])) if isinstance(image_data, dict) else 0
        log(f"文生图提示词 JSON: {image_path}")
    else:
        errors.append(f"模块一 JSON 提取失败：{image_error}")

    if video_rows:
        video_path = md_path.with_name(f"{md_path.stem}_video_prompts.csv")
        reference_image_id = reference_image_id_from_payload(payload, md_path)
        write_video_prompt_csv(video_path, video_fieldnames, video_rows, reference_image_id)
        outputs["video_prompts_csv_path"] = str(video_path)
        outputs["video_prompt_count"] = len(video_rows)
        outputs["video_prompt_reference_image"] = reference_image_id
        log(f"视频片段提示词 CSV: {video_path}")
    else:
        errors.append(f"模块二 CSV 提取失败：{video_error}")

    payload["structured_outputs"] = outputs
    if errors:
        payload["structured_output_errors"] = errors
    return errors


def split_script_into_segments(text, max_segments=12):
    if not text:
        return []
    blocks = []
    current = []
    for line in text.splitlines():
        if line.strip().startswith("镜头 ") and current:
            blocks.append("\n".join(current).strip())
            current = [line]
        else:
            current.append(line)
    if current:
        blocks.append("\n".join(current).strip())
    if len(blocks) <= 1:
        paragraphs = [part.strip() for part in text.split("\n\n") if part.strip()]
        blocks = paragraphs or [text.strip()]
    return blocks[:max_segments]


def build_local_adaptation_scaffold(input_path, source_text, target_model, segment_seconds, notes):
    segments = []
    for index, block in enumerate(split_script_into_segments(source_text), start=1):
        excerpt = block[:900]
        segments.append(
            {
                "segment_id": index,
                "duration_limit_seconds": segment_seconds,
                "source_excerpt": excerpt,
                "video_prompt_draft": f"将原脚本第 {index} 段改写为适合 {target_model} 生成的 {segment_seconds} 秒以内视频片段，保留动作、场景、产品展示与情绪节奏。",
                "first_frame_image_prompt": f"第 {index} 段首帧图：竖屏 9:16，真实 TikTok 原生感画面，人物/场景/产品状态与该片段开头动作一致，主体清晰，[product] 可见。",
            }
        )
    if not segments:
        segments.append(
            {
                "segment_id": 1,
                "duration_limit_seconds": segment_seconds,
                "source_excerpt": "",
                "video_prompt_draft": f"待填入成品脚本后，改写为适合 {target_model} 的 {segment_seconds} 秒以内视频片段。",
                "first_frame_image_prompt": "待填入成品脚本后，生成该片段首帧图描述。",
            }
        )

    lines = [
        "# 脚本适配结果",
        "",
        f"- 输入脚本：{input_path or '未选择'}",
        f"- 目标视频模型：{target_model}",
        f"- 单片段时长上限：{segment_seconds}s",
    ]
    if notes:
        lines.extend(["", "## 适配备注", notes])
    lines.extend(
        [
            "",
            "## 运行说明",
            "当前缺少成品脚本或脚本适配提示词，已生成本地占位框架。请补齐后重新运行，系统会输出可复制给大模型的完整适配提示词包。",
        ]
    )
    lines.extend(["", "## 片段适配框架"])
    for segment in segments:
        lines.extend(
            [
                "",
                f"### 片段 {segment['segment_id']}（≤ {segment_seconds}s）",
                "",
                "**原脚本参考：**",
                "",
                segment["source_excerpt"] or "待补充",
                "",
                "**视频生成提示词草案：**",
                "",
                segment["video_prompt_draft"],
                "",
                "**首帧图描述：**",
                "",
                segment["first_frame_image_prompt"],
            ]
        )
    return "\n".join(lines), segments


def build_adaptation_prompt(config, source_text, target_model, segment_seconds, notes):
    prompt_template = get_script_adaptation_prompt(config)
    target_key = str(target_model).lower()
    is_markdown_segment_model = target_key in {"omni", "grok"}
    source_for_prompt = sanitize_omni_script_text(source_text) if is_markdown_segment_model else source_text
    segment_count_note = ""
    duration = omni_source_duration_seconds(source_for_prompt)
    if target_key == "omni":
        expected_count, duration = expected_omni_segment_count(source_for_prompt, segment_seconds)
        if expected_count is not None:
            segment_count_note = (
                f"\n{str(target_model).title()} 分段数量硬性要求：\n"
                f"原视频总时长约 {duration:.3f}s，单段上限 {segment_seconds}s，"
                f"本次最多输出 {expected_count} 个 Segment，可以少于该数量，但不能超过。\n"
                "不得因为动作、卖点、镜头或场景较多而超过这个最大段数。\n"
            )
    elif target_key == "grok" and duration is not None:
        segment_count_note = (
            "\nGrok 分段时长硬性要求：\n"
            f"原视频总时长约 {duration:.3f}s；本次用户选择的 Grok 单片段时长上限是 {segment_seconds}s。\n"
            f"每个 Segment 必须从 00:00.000 开始，单段时长不能超过 {segment_seconds}s；除原片总长不足 6s 外，单段不能低于 6s。\n"
            "所有 Segment 标题时长相加必须接近原视频总时长，不能明显放大总时长。\n"
            f"如果原片总时长不超过 {segment_seconds}s，优先输出 1 个 Segment；超过 {segment_seconds}s 时按该上限自然拆分，优先使用更少、更完整的长片段。\n"
        )
    padding_note = fixed_duration_padding_note(source_for_prompt, target_model, segment_seconds)
    return f"""{prompt_template}

---

# 系统自动注入变量

目标视频生成模型：
{target_model}

单片段时长上限：
{segment_seconds}s

适配备注：
{notes or "无"}
{segment_count_note}
{padding_note}

---

# 成品脚本内容

{source_for_prompt}
"""


def build_text_payload(prompt, max_output_tokens):
    return {
        "contents": [
            {
                "role": "user",
                "parts": [{"text": prompt}],
            }
        ],
        "generationConfig": {
            "temperature": 0.45,
            "maxOutputTokens": max_output_tokens,
        },
    }


def is_openai_compatible_text_api(base_url, model):
    text = f"{base_url} {model}".lower()
    return "deepseek" in text or "/chat/completions" in text


def build_openai_text_payload(prompt, model, max_output_tokens):
    return {
        "model": model,
        "messages": [
            {
                "role": "user",
                "content": prompt,
            }
        ],
        "temperature": 0.45,
        "max_tokens": max_output_tokens,
    }


def extract_openai_text(response):
    texts = []
    for choice in response.get("choices", []):
        message = choice.get("message") or {}
        content = message.get("content")
        if content:
            texts.append(str(content))
        text = choice.get("text")
        if text:
            texts.append(str(text))
    if texts:
        return "\n".join(texts)
    return extract_text(response)


def run_text_model(prompt, config, label):
    api_key = get_api_key(config)
    if not api_key:
        raise SystemExit("缺少 API Key：脚本适配需要调用文本模型，请先在视频拆解页保存 ModelMesh API Key，或设置 MODELMESH_API_KEY")

    model = (
        str(config.get("script_adaptation_text_model") or "").strip()
        or str(config.get("video_analysis_model") or "").strip()
        or DEFAULT_MODEL
    )
    base_url = str(config.get("modelmesh_base_url") or DEFAULT_BASE_URL).strip()
    max_output_tokens = int(config.get("video_analysis_max_output_tokens", 32768) or 32768)
    openai_compatible = is_openai_compatible_text_api(base_url, model)
    headers = {"Content-Type": "application/json"}
    if openai_compatible:
        headers["Authorization"] = f"Bearer {api_key}"
    else:
        headers["x-goog-api-key"] = api_key

    log(f"开始调用文本模型完成{label}")
    log(f"文本模型: {model}")
    if openai_compatible:
        url = f"{base_url.rstrip('/')}/chat/completions"
        log(f"接口: {url}")
        payload = build_openai_text_payload(prompt, model, max_output_tokens)
        status, response = post_json(url, headers, payload, 240)
        if 200 <= status < 300:
            return extract_openai_text(response), response, "openai-chat-completions"
        message = response.get("error") if isinstance(response, dict) else response
        log(f"  未成功，HTTP {status}: {str(message)[:220]}")
        raise RuntimeError(f"{label}模型调用失败: {json.dumps({'status': status, 'response': response, 'endpoint_style': 'openai-chat-completions'}, ensure_ascii=False)[:1200]}")

    log(f"接口: {base_url.rstrip('/')}/v1beta/models/...:generateContent")

    last_error = None
    for url, endpoint_style in endpoint_variants(base_url, model):
        log(f"尝试接口格式: {endpoint_style}")
        payload = build_text_payload(prompt, max_output_tokens)
        status, response = post_json(url, headers, payload, 240)
        if 200 <= status < 300:
            return extract_text(response), response, endpoint_style
        last_error = {"status": status, "response": response, "endpoint_style": endpoint_style}
        message = response.get("error") if isinstance(response, dict) else response
        log(f"  未成功，HTTP {status}: {str(message)[:220]}")
        time.sleep(0.5)

    raise RuntimeError(f"{label}模型调用失败: {json.dumps(last_error, ensure_ascii=False)[:1200]}")


def run_adapt(config):
    input_path = resolve_path(config.get("script_adaptation_input_path"))
    source_text = read_text(input_path)
    target_model = str(config.get("script_adaptation_target_model") or "veo").strip()
    segment_seconds = int(config.get("script_adaptation_segment_seconds") or 8)
    notes = str(config.get("script_adaptation_notes") or "").strip()
    prompt_template = get_script_adaptation_prompt(config)

    log("开始脚本适配")
    log(f"目标视频生成模型: {target_model}")
    log(f"单片段时长上限: {segment_seconds}s")
    if input_path:
        log(f"输入脚本: {input_path}")
    else:
        log("未选择输入脚本，将生成空白适配框架")

    configured_stem = str(config.get("script_adaptation_output_stem") or "").strip()
    if configured_stem:
        stem = safe_name(configured_stem)
    else:
        stem = f"{timestamp()}_{safe_name(input_path.stem if input_path else 'script_adaptation')}_{safe_name(target_model)}"
    payload = {
        "stage": "script_adaptation",
        "input_path": str(input_path) if input_path else "",
        "output_stem": stem,
        "target_model": target_model,
        "segment_seconds": segment_seconds,
    }

    if source_text and prompt_template:
        adaptation_prompt = build_adaptation_prompt(config, source_text, target_model, segment_seconds, notes)
        adapted_text, raw_response, endpoint_style = run_text_model(adaptation_prompt, config, "脚本适配")
        segments = split_script_into_segments(source_text, max_segments=80)
        if target_model.lower() in {"omni", "grok"}:
            markdown = sanitize_omni_output_markdown(strip_outer_markdown_fence(adapted_text)) + "\n"
            if target_model.lower() == "omni":
                markdown = normalize_omni_machine_structure(
                    markdown,
                    source_text,
                    segment_seconds,
                ) + "\n"
        else:
            markdown = build_clean_adaptation_markdown(
                input_path,
                target_model,
                segment_seconds,
                endpoint_style,
                adapted_text,
            )
        payload.update(
            {
                "adapted_text": adapted_text,
                "raw_response": raw_response,
                "adaptation_prompt": adaptation_prompt,
                "segment_count": len(segments),
                "segments": segments,
            }
        )
        write_outputs("adapt", stem, markdown, payload, config, write_json=False)
        log("脚本适配完成，已生成完整 Markdown")
        return

    if not source_text:
        log("未读取到成品脚本文本，回退为本地框架")
    if not prompt_template:
        log("未读取到脚本适配提示词，回退为本地框架")

    markdown, segments = build_local_adaptation_scaffold(input_path, source_text, target_model, segment_seconds, notes)
    payload["segments"] = segments

    write_outputs("adapt", stem, markdown, payload, config, write_json=False)
    log("脚本适配框架完成")


def run_assemble(config):
    input_dir = resolve_path(config.get("clip_assembly_input_dir"))
    output_name = safe_name(config.get("clip_assembly_output_name") or "assembled_video")
    notes = str(config.get("clip_assembly_notes") or "").strip()
    output_dir = output_dir_for_stage("assemble", config, input_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    log("开始视频生成流程框架")
    log("当前未接入 Veo/可灵等视频生成模型，也不是完整自动剪辑链路；仅处理已有片段目录、生成清单，必要时尝试本地合并。")
    log(f"片段目录: {input_dir or '未选择'}")
    clips = []
    if input_dir and input_dir.exists() and input_dir.is_dir():
        clips = sorted([p for p in input_dir.iterdir() if p.suffix.lower() in {".mp4", ".mov", ".m4v"}])
    log(f"检测到视频片段数量: {len(clips)}")

    manifest = {
        "stage": "clip_assembly",
        "input_dir": str(input_dir) if input_dir else "",
        "clips": [str(path) for path in clips],
        "notes": notes,
    }
    stem = f"{timestamp()}_{output_name}"
    manifest_path = output_dir / f"{stem}_manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    log(f"视频生成清单: {manifest_path}")

    ffmpeg = shutil.which("ffmpeg")
    if clips and ffmpeg:
        concat_list = output_dir / f"{stem}_concat.txt"
        concat_list.write_text("".join(f"file '{path.as_posix()}'\n" for path in clips), encoding="utf-8")
        output_video = output_dir / f"{stem}.mp4"
        log("检测到 ffmpeg，开始尝试无转码合并...")
        result = subprocess.run(
            [ffmpeg, "-y", "-f", "concat", "-safe", "0", "-i", str(concat_list), "-c", "copy", str(output_video)],
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
        )
        if result.returncode == 0 and output_video.exists():
            log(f"已有片段辅助合并完成: {output_video}")
            return
        log("无转码生成未成功，已保留视频生成清单，后续可改为转码合并。")
        log(result.stdout[-1200:])

    markdown = "\n".join(
        [
            "# 视频生成流程框架",
            "",
            "- 状态：流程框架，未接入实际视频生成模型和完整自动剪辑链路",
            f"- 片段目录：{input_dir or '未选择'}",
            f"- 检测片段：{len(clips)} 个",
            f"- 输出名称：{output_name}",
            f"- ffmpeg：{'已检测到' if ffmpeg else '未检测到'}",
            "",
            "## 片段顺序",
            *(f"{index}. {path.name}" for index, path in enumerate(clips, start=1)),
            "",
            "## 备注",
            notes or "待补充",
        ]
    )
    plan_path = output_dir / f"{stem}_plan.md"
    plan_path.write_text(markdown.rstrip() + "\n", encoding="utf-8")
    log(f"视频生成计划: {plan_path}")
    log("视频生成流程框架完成")


def run_publish(config):
    input_path = resolve_path(config.get("video_publish_input_path"))
    account = str(config.get("video_publish_account") or "").strip()
    caption = str(config.get("video_publish_caption") or "").strip()
    tags = str(config.get("video_publish_tags") or "").strip()
    mode = str(config.get("video_publish_mode") or "manual_record").strip()

    log("开始视频发布流程框架")
    log("当前未接入 TikTok 自动登录、账号授权和自动发布；仅生成本地发布计划/记录。")
    log(f"发布模式: {mode}")
    log(f"TikTok账号: {account or '未填写'}")
    log(f"视频文件: {input_path or '未选择'}")
    markdown = "\n".join(
        [
            "# 视频发布流程框架",
            "",
            f"- 状态：流程框架，未接入 TikTok 自动发布",
            f"- 发布状态：待发布",
            f"- 发布模式：{mode}",
            f"- TikTok账号：{account or '未填写'}",
            f"- 视频文件：{input_path or '未选择'}",
            "",
            "## 标题 / 文案",
            caption or "待填写",
            "",
            "## 标签",
            tags or "待填写",
            "",
            "## 接入说明",
            "当前为发布计划/记录框架，尚未接入 TikTok 自动发布。后续确认账号授权方式后，再改为自动发布任务。",
        ]
    )
    payload = {
        "stage": "video_publish",
        "status": "draft",
        "mode": mode,
        "account": account,
        "video_path": str(input_path) if input_path else "",
        "caption": caption,
        "tags": tags,
    }
    write_outputs("publish", f"{timestamp()}_{safe_name(account or 'publish_plan')}", markdown, payload, config)
    log("视频发布流程框架完成")


def parse_numeric(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").replace("%", "").strip())
    except ValueError:
        return None


def normalize_header(value):
    return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())


def find_column(headers, aliases):
    normalized = {normalize_header(header): header for header in headers}
    for alias in aliases:
        match = normalized.get(normalize_header(alias))
        if match:
            return match
    return ""


def normalize_video_id(value):
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    text = str(value).strip()
    if not text or text.upper() in {"N/A", "NA", "NONE", "-", "NULL"}:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def summarize_rows(rows, exclude_keys=None):
    exclude_tokens = [normalize_header(key) for key in (exclude_keys or [])]
    numeric_summary = {}
    if not rows:
        return numeric_summary
    keys = list(rows[0].keys())
    for key in keys:
        normalized_key = normalize_header(key)
        if any(token and token in normalized_key for token in exclude_tokens):
            continue
        values = [parse_numeric(row.get(key, "")) for row in rows]
        values = [value for value in values if value is not None]
        if values:
            numeric_summary[key] = {
                "count": len(values),
                "avg": round(sum(values) / len(values), 4),
                "sum": round(sum(values), 4),
            }
    return numeric_summary


def read_csv_rows(path):
    rows = []
    with path.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            rows.append(row)
    return rows


def read_xlsx_rows(path):
    from openpyxl import load_workbook

    rows = []
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    try:
        headers = next(iterator)
    except StopIteration:
        return rows, {}

    clean_headers = []
    for index, header in enumerate(headers, start=1):
        text = str(header or "").strip()
        clean_headers.append(text or f"column_{index}")

    for values in iterator:
        row = {}
        for header, value in zip(clean_headers, values):
            if hasattr(value, "isoformat"):
                row[header] = value.isoformat(sep=" ")
            elif value is None:
                row[header] = ""
            else:
                row[header] = value
        if any(str(value).strip() for value in row.values()):
            rows.append(row)
    workbook.close()
    return rows


def read_table_rows(path):
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path), "CSV"
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_rows(path), "Excel"
    return [], "Unsupported"


def summarize_table(path):
    rows, table_type = read_table_rows(path)
    return (rows, summarize_rows(rows)), table_type


NATURAL_VIDEO_ID_ALIASES = ["作品ID", "作品id", "视频ID", "视频id", "Video ID", "video id", "aweme id"]
ADS_VIDEO_ID_ALIASES = ["Video ID", "video id", "vidoe id", "视频ID", "视频id", "作品ID", "作品id"]
NATURAL_SIGNAL_ALIASES = ["播放量", "点赞量", "评论量", "账号用户名", "网页地址"]
ADS_SIGNAL_ALIASES = ["Cost", "SKU orders", "Gross revenue", "ROI", "Product ad impressions", "Product ad clicks"]

NATURAL_OUTPUT_COLUMNS = {
    "natural_user_id": ["用户ID", "User ID"],
    "natural_uid": ["UID"],
    "creator_name": ["账号用户名", "TikTok account", "达人", "creator"],
    "caption": ["文案", "Video title", "标题"],
    "cover_url": ["封面", "Cover"],
    "play_url": ["播放地址", "Play URL"],
    "tiktok_url": ["网页地址", "TikTok URL", "URL"],
    "organic_views": ["播放量", "Views"],
    "organic_likes": ["点赞量", "Likes"],
    "organic_comments": ["评论量", "Comments"],
    "created_at": ["创建时间", "发布时间", "Time posted"],
    "updated_at": ["更新时间"],
}

ADS_TEXT_COLUMNS = {
    "ad_video_title": ["Video title", "视频标题", "标题"],
    "ad_tiktok_account": ["TikTok account", "账号用户名", "达人"],
    "ad_creative_type": ["Creative type"],
    "ad_video_source": ["Video source"],
    "ad_status": ["Status"],
    "ad_time_posted": ["Time posted"],
    "currency": ["Currency"],
}

ADS_NUMERIC_SUM_COLUMNS = {
    "ad_cost": ["Cost", "Spend", "广告消耗", "花费"],
    "ad_sku_orders": ["SKU orders", "Orders", "订单", "成交"],
    "ad_gross_revenue": ["Gross revenue", "GMV", "Revenue", "销售额"],
    "ad_impressions": ["Product ad impressions", "Impressions", "曝光"],
    "ad_clicks": ["Product ad clicks", "Clicks", "点击"],
}

ADS_RATE_COLUMNS = {
    "ad_cost_per_order": ["Cost per order"],
    "ad_roi": ["ROI", "ROAS"],
    "ad_click_rate": ["Product ad click rate", "CTR"],
    "ad_conversion_rate": ["Ad conversion rate", "CVR"],
    "ad_2s_view_rate": ["2-second ad video view rate"],
    "ad_6s_view_rate": ["6-second ad video view rate"],
    "ad_25_view_rate": ["25% ad video view rate"],
    "ad_50_view_rate": ["50% ad video view rate"],
    "ad_75_view_rate": ["75% ad video view rate"],
    "ad_100_view_rate": ["100% ad video view rate"],
}

MERGED_METRICS_COLUMNS = [
    "video_id",
    "match_status",
    "creator_name",
    "caption",
    "tiktok_url",
    "play_url",
    "cover_url",
    "created_at",
    "updated_at",
    "organic_views",
    "organic_likes",
    "organic_comments",
    "ad_video_title",
    "ad_tiktok_account",
    "ad_creative_type",
    "ad_video_source",
    "ad_status",
    "ad_time_posted",
    "ad_cost",
    "ad_sku_orders",
    "ad_cost_per_order",
    "ad_gross_revenue",
    "ad_roi",
    "ad_impressions",
    "ad_clicks",
    "ad_click_rate",
    "ad_conversion_rate",
    "ad_2s_view_rate",
    "ad_6s_view_rate",
    "ad_25_view_rate",
    "ad_50_view_rate",
    "ad_75_view_rate",
    "ad_100_view_rate",
    "currency",
    "natural_user_id",
    "natural_uid",
]

MERGED_METRICS_COLUMN_LABELS = {
    "video_id": "视频ID",
    "match_status": "匹配状态",
    "creator_name": "创作者名称",
    "caption": "视频文案",
    "tiktok_url": "TikTok链接",
    "play_url": "播放地址",
    "cover_url": "封面地址",
    "created_at": "作品创建时间",
    "updated_at": "作品更新时间",
    "organic_views": "自然流累计播放量",
    "organic_likes": "自然流累计点赞量",
    "organic_comments": "自然流累计评论量",
    "ad_video_title": "广告视频标题",
    "ad_tiktok_account": "广告账号",
    "ad_creative_type": "广告创意类型",
    "ad_video_source": "广告视频来源",
    "ad_status": "广告状态",
    "ad_time_posted": "广告发布时间",
    "ad_cost": "广告消耗",
    "ad_sku_orders": "广告SKU订单数",
    "ad_cost_per_order": "广告单均成本",
    "ad_gross_revenue": "广告销售额",
    "ad_roi": "广告ROI/ROAS",
    "ad_impressions": "广告曝光量",
    "ad_clicks": "广告点击量",
    "ad_click_rate": "广告点击率",
    "ad_conversion_rate": "广告转化率",
    "ad_2s_view_rate": "2秒广告视频观看率",
    "ad_6s_view_rate": "6秒广告视频观看率",
    "ad_25_view_rate": "25%广告视频观看率",
    "ad_50_view_rate": "50%广告视频观看率",
    "ad_75_view_rate": "75%广告视频观看率",
    "ad_100_view_rate": "100%广告视频观看率",
    "currency": "币种",
    "natural_user_id": "自然流用户ID",
    "natural_uid": "自然流UID",
}

MATCH_STATUS_LABELS = {
    "matched": "两表已匹配",
    "natural_only": "仅自然流存在",
    "ads_only": "仅广告存在",
}


def detect_metrics_table_kind(rows):
    if not rows:
        return ""
    headers = list(rows[0].keys())
    natural_id = find_column(headers, NATURAL_VIDEO_ID_ALIASES)
    ads_id = find_column(headers, ADS_VIDEO_ID_ALIASES)
    natural_signals = sum(1 for alias in NATURAL_SIGNAL_ALIASES if find_column(headers, [alias]))
    ads_signals = sum(1 for alias in ADS_SIGNAL_ALIASES if find_column(headers, [alias]))
    if natural_id and natural_signals >= 2:
        return "natural"
    if ads_id and ads_signals >= 2:
        return "ads"
    return ""


def first_value(row, aliases):
    key = find_column(list(row.keys()), aliases)
    return row.get(key, "") if key else ""


def collect_unique_text(values):
    seen = []
    for value in values:
        text = str(value or "").strip()
        if text and text not in seen:
            seen.append(text)
    return " | ".join(seen[:5])


def average_numeric(values):
    parsed = [parse_numeric(value) for value in values]
    parsed = [value for value in parsed if value is not None]
    if not parsed:
        return ""
    return round(sum(parsed) / len(parsed), 6)


def aggregate_ads_rows(rows, ads_id_column):
    grouped = {}
    skipped_without_video_id = 0
    for row in rows:
        video_id = normalize_video_id(row.get(ads_id_column))
        if not video_id:
            skipped_without_video_id += 1
            continue
        grouped.setdefault(video_id, []).append(row)

    aggregated = {}
    for video_id, group_rows in grouped.items():
        item = {"video_id": video_id, "ad_row_count": len(group_rows)}
        for output_key, aliases in ADS_TEXT_COLUMNS.items():
            item[output_key] = collect_unique_text(first_value(row, aliases) for row in group_rows)
        for output_key, aliases in ADS_NUMERIC_SUM_COLUMNS.items():
            values = [parse_numeric(first_value(row, aliases)) for row in group_rows]
            values = [value for value in values if value is not None]
            item[output_key] = round(sum(values), 6) if values else ""
        for output_key, aliases in ADS_RATE_COLUMNS.items():
            item[output_key] = average_numeric(first_value(row, aliases) for row in group_rows)

        cost = parse_numeric(item.get("ad_cost"))
        orders = parse_numeric(item.get("ad_sku_orders"))
        revenue = parse_numeric(item.get("ad_gross_revenue"))
        impressions = parse_numeric(item.get("ad_impressions"))
        clicks = parse_numeric(item.get("ad_clicks"))
        if cost is not None and orders:
            item["ad_cost_per_order"] = round(cost / orders, 6)
        if cost:
            item["ad_roi"] = round((revenue or 0) / cost, 6)
        if impressions:
            item["ad_click_rate"] = round((clicks or 0) / impressions, 6)
        aggregated[video_id] = item
    return aggregated, skipped_without_video_id


def build_natural_output_row(row):
    item = {}
    for output_key, aliases in NATURAL_OUTPUT_COLUMNS.items():
        item[output_key] = first_value(row, aliases)
    return item


def merge_natural_and_ads(natural_rows, ads_rows):
    natural_headers = list(natural_rows[0].keys()) if natural_rows else []
    ads_headers = list(ads_rows[0].keys()) if ads_rows else []
    natural_id_column = find_column(natural_headers, NATURAL_VIDEO_ID_ALIASES)
    ads_id_column = find_column(ads_headers, ADS_VIDEO_ID_ALIASES)
    if not natural_id_column:
        raise ValueError("自然流数据缺少作品ID字段")
    if not ads_id_column:
        raise ValueError("投放数据缺少 Video ID 字段")

    ads_by_video_id, skipped_ads_without_video_id = aggregate_ads_rows(ads_rows, ads_id_column)
    natural_by_video_id = {}
    natural_order = []
    skipped_natural_without_video_id = 0
    for row in natural_rows:
        video_id = normalize_video_id(row.get(natural_id_column))
        if not video_id:
            skipped_natural_without_video_id += 1
            continue
        if video_id not in natural_by_video_id:
            natural_order.append(video_id)
        natural_by_video_id.setdefault(video_id, []).append(row)

    merged_rows = []
    matched = natural_only = ads_only = 0
    for video_id in natural_order:
        source_rows = natural_by_video_id[video_id]
        natural_item = build_natural_output_row(source_rows[0])
        ads_item = ads_by_video_id.get(video_id, {})
        match_status = "matched" if ads_item else "natural_only"
        matched += 1 if ads_item else 0
        natural_only += 0 if ads_item else 1
        merged_rows.append(
            {
                **{key: "" for key in MERGED_METRICS_COLUMNS},
                "video_id": video_id,
                "match_status": match_status,
                **natural_item,
                **ads_item,
            }
        )

    for video_id, ads_item in ads_by_video_id.items():
        if video_id in natural_by_video_id:
            continue
        ads_only += 1
        merged_rows.append(
            {
                **{key: "" for key in MERGED_METRICS_COLUMNS},
                "video_id": video_id,
                "match_status": "ads_only",
                **ads_item,
            }
        )

    summary = {
        "natural_rows": len(natural_rows),
        "ads_rows": len(ads_rows),
        "merged_rows": len(merged_rows),
        "matched_video_count": matched,
        "natural_only_video_count": natural_only,
        "ads_only_video_count": ads_only,
        "skipped_natural_without_video_id": skipped_natural_without_video_id,
        "skipped_ads_without_video_id": skipped_ads_without_video_id,
    }
    return merged_rows, summary


def write_metrics_table_outputs(stem, rows, config):
    output_dir = product_report_dir("data_attribution", config)
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / f"{stem}.csv"
    output_headers = [MERGED_METRICS_COLUMN_LABELS.get(column, column) for column in MERGED_METRICS_COLUMNS]
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=output_headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            output_row = {}
            for column in MERGED_METRICS_COLUMNS:
                label = MERGED_METRICS_COLUMN_LABELS.get(column, column)
                value = row.get(column, "")
                if column == "match_status":
                    value = MATCH_STATUS_LABELS.get(str(value), value)
                output_row[label] = value
            writer.writerow(output_row)
    log(f"归因 CSV 输出: {csv_path}")
    return csv_path


def localize_summary_keys(summary):
    localized = {}
    for key, value in summary.items():
        localized[MERGED_METRICS_COLUMN_LABELS.get(key, key)] = value
    return localized


def newest_metrics_table_in_dir(path, kind=""):
    if not path or not path.exists() or not path.is_dir():
        return None
    candidates = sorted(
        [item for item in path.rglob("*") if item.is_file() and item.suffix.lower() in METRICS_TABLE_SUFFIXES],
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    if not kind:
        return candidates[0] if candidates else None
    for candidate in candidates:
        try:
            rows, _ = read_table_rows(candidate)
            if detect_metrics_table_kind(rows) == kind:
                return candidate
        except Exception:
            continue
    return None


def resolve_metrics_table_path(config, config_key, kind="", fallback_dir=None):
    table_path = resolve_path(config.get(config_key))
    if table_path and table_path.is_dir():
        return newest_metrics_table_in_dir(table_path, kind)
    if table_path and table_path.exists():
        return table_path
    if fallback_dir:
        return newest_metrics_table_in_dir(fallback_dir, kind)
    return None


def normalize_entrypoint(value, default_value):
    text = str(value or "").strip()
    legacy_leaf = Path(text).name
    if legacy_leaf == "download_natural_flow_data.py":
        return "opc_engine.features.data_attribution.download_natural_flow_data"
    if legacy_leaf == "download_ad_performance_data.py":
        return "opc_engine.features.data_attribution.download_ad_performance_data"
    return text or default_value


def command_for_entrypoint(entrypoint):
    entrypoint = str(entrypoint or "").strip()
    if not entrypoint:
        return None
    module_name = entrypoint.removeprefix("module:").strip()
    if re.fullmatch(r"[A-Za-z_]\w*(\.[A-Za-z_]\w*)+", module_name):
        return [sys.executable, "-m", module_name]
    script_path = resolve_path(entrypoint)
    if not script_path or not script_path.exists() or not script_path.is_file():
        raise FileNotFoundError(f"内部下载入口不存在: {entrypoint}")
    suffix = script_path.suffix.lower()
    if suffix == ".py":
        return [sys.executable, str(script_path)]
    if suffix in {".sh", ".bash"}:
        return ["bash", str(script_path)]
    if os.access(script_path, os.X_OK):
        return [str(script_path)]
    return ["bash", str(script_path)]


def sanitize_download_log(text):
    replacements = {
        "GMVMax": "投放数据平台",
        "GMV Max": "投放数据平台",
        "TikTok Ads": "投放数据平台",
    }
    cleaned = str(text)
    for old, new in replacements.items():
        cleaned = cleaned.replace(old, new)
    cleaned = re.sub(r"https?://ads\.tiktok\.com[^\s)]+", "投放数据页面", cleaned)
    return cleaned


def run_download_script(config, label, script_key, default_entrypoint, output_subdir="", extra_env=None):
    entrypoint = normalize_entrypoint(config.get(script_key), default_entrypoint)
    output_dir = raw_data_dir(output_subdir or "downloads", config)
    notes = str(config.get("data_attribution_download_notes") or "").strip()
    output_dir.mkdir(parents=True, exist_ok=True)

    log(f"开始下载{label}")
    log(f"保存目录: {output_dir}")

    if not entrypoint:
        markdown = "\n".join(
            [
                f"# {label}下载",
                "",
                "- 状态：未配置内部下载入口",
                f"- 保存目录：{output_dir}",
                "",
                "## 下一步",
                "配置本地下载入口后再运行。",
                "",
                "## 备注",
                notes or "待补充",
            ]
        )
        write_outputs(
            "metrics",
            f"{timestamp()}_data_attribution_download_plan",
            markdown,
            {
                "stage": "data_attribution_download",
                "status": "script_not_configured",
                "label": label,
                "output_dir": str(output_dir),
                "notes": notes,
            },
            config,
        )
        log("未配置内部下载入口，已生成阶段一流程清单")
        return

    command = command_for_entrypoint(entrypoint)
    env = os.environ.copy()
    env["DATA_ATTRIBUTION_OUTPUT_DIR"] = str(output_dir)
    env["OPC_DATA_ATTRIBUTION_OUTPUT_DIR"] = str(output_dir)
    for key, value in (extra_env or {}).items():
        if value is not None and str(value).strip():
            env[key] = str(value)
    natural_url = config.get("natural_flow_management_url")
    natural_login_url = config.get("natural_flow_login_url")
    natural_group = config.get("natural_flow_account_group")
    natural_export_text = config.get("natural_flow_export_button_text_re")
    if natural_url:
        env["NATURAL_FLOW_MANAGEMENT_URL"] = str(natural_url)
    if natural_login_url:
        env["NATURAL_FLOW_LOGIN_URL"] = str(natural_login_url)
    if natural_group:
        env["NATURAL_FLOW_ACCOUNT_GROUP"] = str(natural_group)
        env["NATURAL_DATA_ACCOUNT_GROUP"] = str(natural_group)
    if natural_export_text:
        env["NATURAL_FLOW_EXPORT_TEXT_RE"] = str(natural_export_text)
        env["NATURAL_DATA_EXPORT_TEXT_RE"] = str(natural_export_text)
    log("执行内部下载流程...")
    process = subprocess.Popen(
        command,
        cwd=str(ROOT),
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        bufsize=1,
    )
    captured = []
    if process.stdout:
        for line in process.stdout:
            text = sanitize_download_log(line.rstrip())
            captured.append(text)
            log(text)
    exit_code = process.wait()
    downloaded_files = sorted([path for path in output_dir.rglob("*") if path.is_file()], key=lambda item: item.stat().st_mtime, reverse=True)

    markdown = "\n".join(
            [
                f"# {label}下载",
                "",
                f"- 状态：{'成功' if exit_code == 0 else '失败'}",
                f"- 退出码：{exit_code}",
                f"- 保存目录：{output_dir}",
                f"- 检测到文件：{len(downloaded_files)} 个",
                "",
                "## 最新文件",
            *(f"- {path}" for path in downloaded_files[:30]),
            "",
            "## 脚本输出摘要",
            "\n".join(captured[-80:]) or "无输出",
            "",
            "## 备注",
            notes or "待补充",
        ]
    )
    write_outputs(
        "metrics",
        f"{timestamp()}_{safe_name(label)}_download",
        markdown,
        {
            "stage": "data_attribution_download",
            "status": "success" if exit_code == 0 else "failed",
            "exit_code": exit_code,
            "label": label,
            "output_dir": str(output_dir),
            "downloaded_files": [str(path) for path in downloaded_files],
            "notes": notes,
        },
        config,
    )
    if exit_code != 0:
        raise SystemExit(exit_code)
    log(f"{label}下载完成")


def run_metrics_natural_download(config):
    run_download_script(
        config,
        "自然流数据",
        "data_attribution_download_script_path",
        "opc_engine.features.data_attribution.download_natural_flow_data",
        output_subdir="natural_flow",
    )


def run_metrics_ads_download(config):
    run_download_script(
        config,
        "投放数据",
        "data_attribution_ads_download_script_path",
        "opc_engine.features.data_attribution.download_ad_performance_data",
        output_subdir="ad_performance",
        extra_env={"AD_PERFORMANCE_DATE_RANGE": "yesterday"},
    )


def run_metrics_download(config):
    run_metrics_natural_download(config)
    run_metrics_ads_download(config)


def run_metrics(config):
    configured_download_dir = str(config.get("data_attribution_download_output_dir") or "").strip()
    if configured_download_dir == "metrics/raw_downloads":
        download_output_dir = product_project_root(config) / "raw_data"
    else:
        download_output_dir = resolve_path(configured_download_dir) or (product_project_root(config) / "raw_data")
    legacy_input_path = resolve_path(config.get("data_recovery_input_path"))
    natural_input_path = resolve_metrics_table_path(
        config,
        "data_recovery_natural_input_path",
        "natural",
        download_output_dir,
    )
    ads_input_path = resolve_metrics_table_path(config, "data_recovery_ads_input_path", "ads", download_output_dir)
    if legacy_input_path and legacy_input_path.is_dir():
        natural_input_path = natural_input_path or newest_metrics_table_in_dir(legacy_input_path, "natural")
        ads_input_path = ads_input_path or newest_metrics_table_in_dir(legacy_input_path, "ads")
        input_path = natural_input_path or ads_input_path or newest_metrics_table_in_dir(legacy_input_path)
    else:
        input_path = legacy_input_path
    manual_metrics = str(config.get("data_recovery_manual_metrics") or "").strip()
    if not input_path and not (natural_input_path and ads_input_path) and download_output_dir:
        input_path = newest_metrics_table_in_dir(download_output_dir)

    log("开始数据归因阶段二：整理与分析")
    log(f"自然流数据: {'已找到最新原始表' if natural_input_path else '未找到'}")
    log(f"投放数据: {'已找到最新原始表' if ads_input_path else '未找到'}")

    if natural_input_path and ads_input_path:
        natural_rows, natural_table_type = read_table_rows(natural_input_path)
        ads_rows, ads_table_type = read_table_rows(ads_input_path)
        log(f"读取自然流 {natural_table_type} 行数: {len(natural_rows)}")
        log(f"读取投放 {ads_table_type} 行数: {len(ads_rows)}")
        merged_rows, attribution_summary = merge_natural_and_ads(natural_rows, ads_rows)
        metrics_summary = summarize_rows(
            merged_rows,
            exclude_keys=["id", "url", "path", "time", "created", "updated", "caption", "title", "status", "type", "currency"],
        )
        localized_metrics_summary = localize_summary_keys(metrics_summary)
        stem = f"{timestamp()}_作品归因汇总"
        csv_path = write_metrics_table_outputs(stem, merged_rows, config)
        markdown = "\n".join(
            [
                "# 数据归因结果",
                "",
                "阶段二已按同一个作品维度，把自然流数据和投放数据合并成一张作品归因表。",
                "",
                f"- 合并 CSV：{csv_path}",
                "",
                "## 匹配结果",
                f"- 自然流原始行数：{attribution_summary['natural_rows']}",
                f"- 投放原始行数：{attribution_summary['ads_rows']}",
                f"- 合并后作品数：{attribution_summary['merged_rows']}",
                f"- 两表匹配作品数：{attribution_summary['matched_video_count']}",
                f"- 仅自然流存在：{attribution_summary['natural_only_video_count']}",
                f"- 仅投放存在：{attribution_summary['ads_only_video_count']}",
                f"- 自然流无作品ID跳过：{attribution_summary['skipped_natural_without_video_id']}",
                f"- 投放无 Video ID 跳过：{attribution_summary['skipped_ads_without_video_id']}",
                "",
                "## 手动数据",
                manual_metrics or "待补充",
                "",
                "## 数值字段汇总",
                json.dumps(localized_metrics_summary, ensure_ascii=False, indent=2) if localized_metrics_summary else "暂无可汇总数值字段",
            ]
        )
        payload = {
            "stage": "data_attribution",
            "mode": "natural_ads_merge",
            "natural_input_detected": bool(natural_input_path),
            "ads_input_detected": bool(ads_input_path),
            "merged_csv_path": str(csv_path),
            "manual_metrics": manual_metrics,
            "attribution_summary": attribution_summary,
            "numeric_summary": localized_metrics_summary,
        }
        write_outputs("metrics", stem, markdown, payload, config)
        log("数据归因阶段二完成")
        return

    log(f"单表数据来源: {'已找到最新原始表' if input_path else '手动填写/未选择'}")

    rows = []
    summary = {}
    table_type = ""
    if input_path and input_path.exists() and input_path.suffix.lower() in METRICS_TABLE_SUFFIXES:
        (rows, summary), table_type = summarize_table(input_path)
        log(f"读取 {table_type} 行数: {len(rows)}")
    markdown = "\n".join(
        [
            "# 数据归因结果",
            "",
            f"- 数据来源：{'已读取最新原始表' if input_path else '手动填写/未选择'}",
            f"- 数据类型：{table_type or '手动/未识别'}",
            f"- 数据行数：{len(rows)}",
            "",
            "## 手动数据",
            manual_metrics or "待补充",
            "",
            "## 数值字段汇总",
            json.dumps(summary, ensure_ascii=False, indent=2) if summary else "暂无可汇总数值字段",
        ]
    )
    payload = {
        "stage": "data_attribution",
        "input_path": str(input_path) if input_path else "",
        "table_type": table_type,
        "manual_metrics": manual_metrics,
        "row_count": len(rows),
        "numeric_summary": summary,
    }
    write_outputs("metrics", f"{timestamp()}_{safe_name(input_path.stem if input_path else 'metrics')}", markdown, payload, config)
    log("数据归因阶段二完成")


def run_optimize(config):
    script_path = resolve_path(config.get("script_optimization_input_path"))
    metrics_path = resolve_path(config.get("script_optimization_metrics_path"))
    notes = str(config.get("script_optimization_notes") or "").strip()
    source_script = read_text(script_path)
    metrics_text = read_text(metrics_path)

    log("开始脚本优化框架")
    log(f"原脚本: {script_path or '未选择'}")
    log(f"数据文件: {metrics_path or '未选择'}")
    markdown = "\n".join(
        [
            "# 脚本优化建议",
            "",
            f"- 原脚本：{script_path or '未选择'}",
            f"- 数据文件：{metrics_path or '未选择'}",
            "",
            "## 加权评估框架",
            "- 播放完成/停留：判断 Hook 与前 3 秒是否成立。",
            "- 点击/互动：判断冲突、痛点、评论诱因是否成立。",
            "- 转化/GMV：判断产品机制、信任背书、价格锚点是否成立。",
            "- 多视频加权平均：后续接入发布数据后，以视频级指标反推脚本表现。",
            "",
            "## 当前优化建议草案",
            "1. 先定位数据最弱的环节：开头停留、互动、点击、成交。",
            "2. 保留表现强的镜头结构，只替换低效话术和弱视觉证据。",
            "3. 对同一脚本拆出 A/B 版本：强冲突版、强证明版、强价格锚点版。",
            "",
            "## 备注",
            notes or "待补充",
            "",
            "## 原脚本摘要",
            source_script[:1600] or "待补充",
            "",
            "## 数据摘要",
            metrics_text[:1600] or "待补充",
        ]
    )
    payload = {
        "stage": "script_optimization",
        "script_path": str(script_path) if script_path else "",
        "metrics_path": str(metrics_path) if metrics_path else "",
        "notes": notes,
    }
    write_outputs("optimize", f"{timestamp()}_{safe_name(script_path.stem if script_path else 'script_optimization')}", markdown, payload, config)
    log("脚本优化框架完成")


def parse_args():
    parser = argparse.ArgumentParser(description="Run local content distribution workflow scaffolds.")
    parser.add_argument(
        "stage",
        choices=["adapt", "assemble", "publish", "metrics_download", "metrics_natural_download", "metrics_ads_download", "metrics", "optimize"],
        help="要运行的工作流阶段",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    stage_inputs = {
        "adapt": ("script_adaptation",),
        "assemble": ("video_generation",),
        "publish": ("video_publish",),
        "metrics_download": ("data_attribution",),
        "metrics_natural_download": ("data_attribution",),
        "metrics_ads_download": ("data_attribution",),
        "metrics": ("data_attribution",),
        "optimize": ("script_optimization",),
    }
    config = load_config(*stage_inputs.get(args.stage, ()))
    require_product_project(config, "执行内容工作流")
    ensure_project_dirs(config)
    {
        "adapt": run_adapt,
        "assemble": run_assemble,
        "publish": run_publish,
        "metrics_download": run_metrics_download,
        "metrics_natural_download": run_metrics_natural_download,
        "metrics_ads_download": run_metrics_ads_download,
        "metrics": run_metrics,
        "optimize": run_optimize,
    }[args.stage](config)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        log(f"任务失败: {exc}")
        sys.exit(1)
